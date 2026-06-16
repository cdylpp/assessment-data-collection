from __future__ import annotations

import argparse
import sys
import tempfile
import unittest
from pathlib import Path

import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = REPO_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from template_generator import (  # noqa: E402
    RosterUidConfig,
    TemplateGenerationRequest,
    build_candidate_uid_from_values,
    generate_excel_template,
    generate_template_workbook,
    load_generation_inputs,
    load_yaml,
    load_roster,
    request_from_namespace,
)
from evaluator import Evaluator  # noqa: E402
from event_resolver import EventResolver  # noqa: E402


class TemplateGeneratorSmokeTest(unittest.TestCase):
    def test_load_generation_inputs(self) -> None:
        request = TemplateGenerationRequest(
            config_path=(REPO_ROOT / "config" / "config.yaml").resolve()
        )

        loaded = load_generation_inputs(request)

        self.assertEqual(loaded.config_doc["registry_name"], "usna_screener_evos")
        self.assertTrue(loaded.metrics_doc["metrics"])
        self.assertTrue(loaded.evolutions_doc["evolutions"])
        self.assertTrue(loaded.roster_rows)
        self.assertEqual(
            loaded.roster_rows[0]["uid"],
            build_candidate_uid_from_values(["Andrew", "Lucas", "2004-12-16"]),
        )
        self.assertIsNotNone(loaded.assessment_config)
        assert loaded.assessment_config is not None
        plan = EventResolver(loaded.assessment_config).resolve()
        self.assertEqual(plan.event_id, "soac_fy_2026_block1")
        self.assertIn(
            "soac_fy_2026_block1__evo_grit_pt__5",
            [instance.instance_id for instance in plan.instances],
        )
        grit_fifth = next(
            instance
            for instance in plan.instances
            if instance.instance_id == "soac_fy_2026_block1__evo_grit_pt__5"
        )
        self.assertEqual(grit_fifth.sheet_name, "Grit PT #5")
        self.assertEqual(grit_fifth.event_occurrence_index, 5)

    def test_load_roster_uses_existing_uid_column_from_config(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            roster_path = Path(temp_dir) / "roster.csv"
            roster_path.write_text(
                "Candidate ID,First,Last,DOB\nabc-123,Jane,Doe,2001-02-03\n",
                encoding="utf-8",
            )

            rows = load_roster(
                roster_path,
                uid_config=RosterUidConfig(
                    mode="existing",
                    source_column="Candidate ID",
                    key_columns=[],
                ),
            )

            self.assertEqual(rows[0]["uid"], "abc-123")
            self.assertEqual(rows[0]["first"], "Jane")
            self.assertEqual(rows[0]["last"], "Doe")

    def test_load_roster_generates_uid_from_configured_key_columns(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            roster_path = Path(temp_dir) / "roster.csv"
            roster_path.write_text(
                "First,Last,DOB\nJane,Doe,02/03/2001\n",
                encoding="utf-8",
            )

            rows = load_roster(
                roster_path,
                uid_config=RosterUidConfig(
                    mode="generated",
                    source_column=None,
                    key_columns=["last", "first", "dob"],
                ),
            )

            self.assertEqual(
                rows[0]["uid"],
                build_candidate_uid_from_values(["Doe", "Jane", "2001-02-03"]),
            )
            self.assertEqual(rows[0]["dob"], "2001-02-03")

    def test_generate_template_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "phase1_smoke.xlsx"
            request = TemplateGenerationRequest(
                config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                output_path=output_path,
                block_number="B01",
                fiscal_year="2026",
                entry_rows=10,
            )

            generated_path = generate_template_workbook(request)

            self.assertEqual(generated_path, output_path.resolve())
            self.assertTrue(generated_path.exists())

            workbook = load_workbook(generated_path)
            try:
                loaded = load_generation_inputs(request)
                self.assertIn("META", workbook.sheetnames)
                self.assertIn("ROSTER", workbook.sheetnames)
                self.assertIn("LOOKUPS", workbook.sheetnames)
                self.assertIn("PST", workbook.sheetnames)
                self.assertIn("Grit PT #1", workbook.sheetnames)
                self.assertIn("Grit PT #5", workbook.sheetnames)
                self.assertIn("Log Carry around O Course", workbook.sheetnames)
                self.assertIn("IBS Land Portage #2", workbook.sheetnames)
                pst = workbook["PST"]
                self.assertFalse(workbook["META"].protection.sheet)
                self.assertFalse(workbook["ROSTER"].protection.sheet)
                self.assertFalse(pst.protection.sheet)
                self.assertFalse(pst.row_dimensions[2].hidden)
                self.assertTrue(pst.row_dimensions[3].hidden)
                roster_fields = loaded.config_doc["sheet_contract"][
                    "locked_left_columns"
                ]
                self.assertTrue(pst.column_dimensions["A"].hidden)
                self.assertEqual(pst.cell(row=3, column=1).value, "uid")
                self.assertEqual(
                    pst.cell(row=4, column=1).value,
                    workbook["ROSTER"]["A2"].value,
                )
                for col_idx, field in enumerate(roster_fields, start=2):
                    self.assertEqual(pst.cell(row=3, column=col_idx).value, field)
                    roster_column = {"uid": "A", "first": "B", "last": "C"}.get(field)
                    if roster_column:
                        self.assertEqual(
                            pst.cell(row=4, column=col_idx).value,
                            workbook["ROSTER"]["{0}2".format(roster_column)].value,
                        )
                metric_start_col = len(roster_fields) + 2
                run_time_col = metric_start_col + 3
                run_time_col_letter = get_column_letter(run_time_col)
                first_metric_cell = pst.cell(row=4, column=metric_start_col)
                self.assertEqual(
                    pst.cell(row=3, column=metric_start_col).value,
                    "m_push_ups",
                )
                if "cohort" in roster_fields:
                    cohort_col = roster_fields.index("cohort") + 2
                    self.assertEqual(pst.cell(row=4, column=cohort_col).value, "A")
                self.assertIsNotNone(first_metric_cell.comment)
                self.assertIn("Candidate: Lucas Andrew", first_metric_cell.comment.text)
                self.assertEqual(pst.freeze_panes, "E4")
                self.assertEqual(
                    pst.cell(row=4, column=run_time_col).number_format,
                    "[mm]:ss",
                )
                timed_metric_columns = []
                for sheet_name in workbook.sheetnames:
                    if sheet_name in {"META", "ROSTER", "LOOKUPS"}:
                        continue
                    ws = workbook[sheet_name]
                    for col_idx in range(1, ws.max_column + 1):
                        metric_id = ws.cell(row=3, column=col_idx).value
                        metric = loaded.metrics_by_id.get(metric_id)
                        if metric and metric.get("type") == "timed":
                            timed_metric_columns.append((sheet_name, col_idx))
                            self.assertEqual(
                                ws.cell(row=4, column=col_idx).number_format,
                                "[mm]:ss",
                            )
                self.assertTrue(timed_metric_columns)
                timed_validations = [
                    validation
                    for validation in pst.data_validations.dataValidation
                    if validation.type == "custom"
                    and validation.sqref
                    and "{0}4:{0}13".format(run_time_col_letter)
                    in str(validation.sqref)
                ]
                self.assertEqual(len(timed_validations), 1)
                self.assertIn(
                    "MOD({0}4,1)<0.6".format(run_time_col_letter),
                    timed_validations[0].formula1,
                )
                self.assertEqual(pst.row_dimensions[4].height, 30)
                self.assertEqual(
                    pst.column_dimensions[get_column_letter(metric_start_col)].width,
                    24,
                )
                self.assertNotEqual(
                    first_metric_cell.fill.fgColor.rgb,
                    pst.cell(row=5, column=metric_start_col).fill.fgColor.rgb,
                )

                grit = workbook["Grit PT #1"]
                grit_config = next(
                    evolution
                    for evolution in loaded.evolutions_doc["evolutions"]
                    if evolution["evolution_id"] == "evo_grit_pt"
                )
                occurrence_count = grit_config["metric_occurrences"][
                    "m_grit_pt_physicality"
                ]
                self.assertEqual(grit["E1"].value, "Grit PT - Physicality")
                self.assertEqual(grit["X1"].value, "evolution_id")
                self.assertEqual(grit["Y1"].value, "evo_grit_pt")
                self.assertEqual(grit["X2"].value, "event_id")
                self.assertEqual(grit["Y2"].value, "soac_fy_2026_block1")
                self.assertEqual(grit["X4"].value, "event_instance_id")
                self.assertEqual(
                    grit["Y4"].value,
                    "soac_fy_2026_block1__evo_grit_pt__1",
                )
                self.assertEqual(grit["X5"].value, "event_occurrence_index")
                self.assertEqual(grit["Y5"].value, 1)
                self.assertEqual(grit["E2"].value, 1)
                self.assertEqual(grit["J2"].value, occurrence_count)
                self.assertEqual(grit["E3"].value, "m_grit_pt_physicality")
                self.assertIn(
                    "E1:J1",
                    [str(merged_range) for merged_range in grit.merged_cells.ranges],
                )
                repeated_fill = grit["E2"].fill.fgColor.rgb
                self.assertNotEqual(grit["E4"].fill.fgColor.rgb, "00FFFFFF")
                self.assertEqual(grit["E4"].fill.fgColor.rgb, grit["J4"].fill.fgColor.rgb)
                self.assertNotEqual(grit["E4"].fill.fgColor.rgb, grit["K4"].fill.fgColor.rgb)
                self.assertNotEqual(grit["E4"].fill.fgColor.rgb, grit["E5"].fill.fgColor.rgb)
                self.assertEqual(grit["E4"].border.right.style, "thin")
                self.assertEqual(grit["E9"].fill.fgColor.rgb, "00FFFFFF")
                for offset in range(occurrence_count):
                    col_idx = 5 + offset
                    self.assertEqual(
                        grit.cell(row=2, column=col_idx).value,
                        offset + 1,
                    )
                    self.assertEqual(
                        grit.cell(row=3, column=col_idx).value,
                        "m_grit_pt_physicality",
                    )
                    self.assertEqual(
                        grit.cell(row=2, column=col_idx).fill.fgColor.rgb,
                        repeated_fill,
                    )
            finally:
                workbook.close()

    def test_cli_events_path_can_add_config_based_definitions(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            config_dir = temp_root / "config"
            event_dir = temp_root / "custom_events"
            config_dir.mkdir()
            event_dir.mkdir()

            metrics_path = config_dir / "extra-metrics.yaml"
            metrics_path.write_text(
                yaml.safe_dump(
                    {
                        "metrics": [
                            {
                                "metric_id": "m_custom_score",
                                "display_name": "Custom Score",
                                "type": "integer",
                                "input_kind": "measured",
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )
            evolutions_path = config_dir / "extra-evolutions.yaml"
            evolutions_path.write_text(
                yaml.safe_dump(
                    {
                        "evolutions": [
                            {
                                "evolution_id": "evo_custom",
                                "display_name": "Custom Evolution",
                                "sheet_name": "Custom Evolution",
                                "metric_ids": ["m_custom_score"],
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )
            events_path = event_dir / "events.yaml"
            events_path.write_text(
                yaml.safe_dump(
                    {
                        "files": {
                            "metrics": "config/extra-metrics.yaml",
                            "evolutions": "config/extra-evolutions.yaml",
                        },
                        "events": [
                            {
                                "id": "custom_event",
                                "name": "Custom Event",
                                "evolutions": [["evo_custom"]],
                            }
                        ],
                    }
                ),
                encoding="utf-8",
            )

            config_doc = load_yaml(REPO_ROOT / "config" / "config.yaml")
            config_doc["files"] = {
                "metrics": str((REPO_ROOT / "config" / "metrics.yaml").resolve()),
                "evolutions": str((REPO_ROOT / "config" / "evolutions.yaml").resolve()),
                "roster": str((REPO_ROOT / "config" / "roster.csv").resolve()),
                "master": str((REPO_ROOT / "config" / "master-config.yaml").resolve()),
            }
            config_path = config_dir / "config.yaml"
            config_path.write_text(yaml.safe_dump(config_doc), encoding="utf-8")
            output_path = temp_root / "custom_event.xlsx"

            args = argparse.Namespace(
                config=str(config_path),
                roster=None,
                output=str(output_path),
                events=str(events_path),
                block_number="B01",
                fiscal_year="2026",
                entry_rows=3,
                event_id="custom_event",
            )
            request = request_from_namespace(args)
            generated_path = generate_template_workbook(request)

            self.assertEqual(request.events_path, events_path.resolve())
            self.assertTrue(generated_path.exists())
            loaded = load_generation_inputs(request)
            self.assertEqual(loaded.events_path, events_path.resolve())
            self.assertIn("m_custom_score", loaded.metrics_by_id)

            workbook = load_workbook(generated_path)
            try:
                self.assertIn("Custom Evolution", workbook.sheetnames)
                self.assertEqual(
                    workbook["Custom Evolution"]["E3"].value,
                    "m_custom_score",
                )
            finally:
                workbook.close()

            generated_by_function = generate_excel_template(
                config_path=config_path,
                output_path=temp_root / "custom_event_function.xlsx",
                events_path=events_path,
                block_number="B01",
                fiscal_year="2026",
                entry_rows=3,
                event_id="custom_event",
            )
            self.assertTrue(generated_by_function.exists())

    def test_evaluator_reports_event_and_metric_inconsistencies(self) -> None:
        result = Evaluator(
            metrics_doc={"metrics": [{"metric_id": "m_defined"}]},
            evolutions_doc={
                "evolutions": [
                    {
                        "evolution_id": "evo_blank",
                        "metric_ids": [],
                    },
                    {
                        "evolution_id": "evo_singular",
                        "metric_id": ["m_defined"],
                    },
                    {
                        "evolution_id": "evo_unknown_metric",
                        "metric_ids": ["m_missing"],
                    },
                ]
            },
            events_doc={
                "events": [
                    {
                        "event_id": "event_invalid",
                        "evolutions": [
                            "evo_blank",
                            {"id": "evo_missing"},
                        ],
                    }
                ]
            },
        ).evaluate()

        messages = [issue.message for issue in result.errors]
        self.assertFalse(result.is_valid)
        self.assertTrue(
            any("evo_missing" in message for message in messages),
            result.error_message(),
        )
        self.assertTrue(
            any("evo_blank" in message for message in messages),
            result.error_message(),
        )
        self.assertTrue(
            any("'metric_id'; expected 'metric_ids'" in message for message in messages),
            result.error_message(),
        )

    def test_candidate_comments_can_be_disabled(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            config_dir = temp_root / "config"
            config_dir.mkdir()
            config_doc = load_yaml(REPO_ROOT / "config" / "config.yaml")
            config_doc["files"] = {
                "metrics": str((REPO_ROOT / "config" / "metrics.yaml").resolve()),
                "evolutions": str((REPO_ROOT / "config" / "evolutions.yaml").resolve()),
                "roster": str((REPO_ROOT / "config" / "roster.csv").resolve()),
                "master": str((REPO_ROOT / "config" / "master-config.yaml").resolve()),
            }
            config_doc["workbook_ui"] = {"candidate_name_comments": False}
            config_path = config_dir / "config.yaml"
            config_path.write_text(yaml.safe_dump(config_doc), encoding="utf-8")
            output_path = temp_root / "comments_disabled.xlsx"

            generated_path = generate_template_workbook(
                TemplateGenerationRequest(
                    config_path=config_path,
                    output_path=output_path,
                    block_number="B01",
                    fiscal_year="2026",
                    entry_rows=10,
                )
            )

            workbook = load_workbook(generated_path)
            try:
                self.assertIsNone(workbook["PST"]["E4"].comment)
            finally:
                workbook.close()

    def test_evolution_can_override_locked_left_columns(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            config_dir = temp_root / "config"
            config_dir.mkdir()
            evolutions_path = config_dir / "evolutions.yaml"
            config_doc = load_yaml(REPO_ROOT / "config" / "config.yaml")
            evolutions_doc = load_yaml(REPO_ROOT / "config" / "evolutions.yaml")
            evolutions_doc["evolutions"][0]["sheet_contract"] = {
                "locked_left_columns": ["uid", "first", "last"]
            }
            evolutions_path.write_text(
                yaml.safe_dump(evolutions_doc), encoding="utf-8"
            )
            config_doc["files"] = {
                "metrics": str((REPO_ROOT / "config" / "metrics.yaml").resolve()),
                "evolutions": str(evolutions_path.resolve()),
                "roster": str((REPO_ROOT / "config" / "roster.csv").resolve()),
                "master": str((REPO_ROOT / "config" / "master-config.yaml").resolve()),
            }
            config_path = config_dir / "config.yaml"
            config_path.write_text(yaml.safe_dump(config_doc), encoding="utf-8")
            output_path = temp_root / "override.xlsx"

            generated_path = generate_template_workbook(
                TemplateGenerationRequest(
                    config_path=config_path,
                    output_path=output_path,
                    block_number="B01",
                    fiscal_year="2026",
                    entry_rows=10,
                )
            )

            workbook = load_workbook(generated_path)
            try:
                self.assertTrue(workbook["PST"].column_dimensions["A"].hidden)
                self.assertEqual(workbook["PST"]["A3"].value, "uid")
                self.assertEqual(workbook["PST"]["B3"].value, "first")
                self.assertEqual(workbook["PST"]["D3"].value, "m_push_ups")
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
