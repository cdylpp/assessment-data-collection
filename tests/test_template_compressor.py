from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import timedelta
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = REPO_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from template_compressor import (  # noqa: E402
    TemplateCompressionConflictError,
    TemplateCompressionRequest,
    compress_template_workbooks,
)
from template_generator import TemplateGenerationRequest, generate_template_workbook  # noqa: E402


def set_sheet_value(path: Path, sheet_name: str, cell_ref: str, value: object) -> None:
    workbook = load_workbook(path)
    try:
        ws = workbook[sheet_name]
        ws[cell_ref] = value
        workbook.save(path)
    finally:
        workbook.close()


class TemplateCompressorTest(unittest.TestCase):
    def test_compress_template_workbooks_copies_sparse_scores(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            node_a_path = temp_root / "node_a.xlsx"
            node_b_path = temp_root / "node_b.xlsx"
            output_path = temp_root / "node.xlsx"

            for path in (node_a_path, node_b_path):
                generate_template_workbook(
                    TemplateGenerationRequest(
                        config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                        output_path=path,
                        block_number="101",
                        fiscal_year="2026",
                        entry_rows=10,
                    )
                )

            set_sheet_value(node_a_path, "PST", "E4", 61)
            set_sheet_value(node_b_path, "PST", "F4", 62)
            set_sheet_value(node_b_path, "Grit PT #1", "E5", 3)

            generated_path = compress_template_workbooks(
                TemplateCompressionRequest(
                    config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                    workbook_paths=[node_a_path, node_b_path],
                    output_path=output_path,
                )
            )

            self.assertEqual(generated_path, output_path.resolve())
            workbook = load_workbook(generated_path, data_only=True)
            try:
                self.assertEqual(workbook["PST"]["E4"].value, 61)
                self.assertEqual(workbook["PST"]["F4"].value, 62)
                self.assertEqual(workbook["Grit PT #1"]["E5"].value, 3)
            finally:
                workbook.close()

    def test_compress_template_workbooks_normalizes_five_mile_mm_ss(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            node_a_path = temp_root / "node_a.xlsx"
            node_b_path = temp_root / "node_b.xlsx"
            output_path = temp_root / "node.xlsx"

            for path in (node_a_path, node_b_path):
                generate_template_workbook(
                    TemplateGenerationRequest(
                        config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                        output_path=path,
                        block_number="101",
                        fiscal_year="2026",
                        entry_rows=10,
                    )
                )

            set_sheet_value(node_a_path, "5 Mile Run", "E4", 23.06)
            set_sheet_value(node_b_path, "5 Mile Run", "E5", 30.57)

            generated_path = compress_template_workbooks(
                TemplateCompressionRequest(
                    config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                    workbook_paths=[node_a_path, node_b_path],
                    output_path=output_path,
                )
            )

            workbook = load_workbook(generated_path, data_only=True)
            try:
                row_four_value = workbook["5 Mile Run"]["E4"].value
                row_five_value = workbook["5 Mile Run"]["E5"].value
                self.assertEqual(row_four_value, timedelta(minutes=23, seconds=6))
                self.assertEqual(row_five_value, timedelta(minutes=30, seconds=57))
            finally:
                workbook.close()

    def test_compress_template_workbooks_reports_conflicting_scores(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            node_a_path = temp_root / "node_a.xlsx"
            node_b_path = temp_root / "node_b.xlsx"
            output_path = temp_root / "node.xlsx"

            for path in (node_a_path, node_b_path):
                generate_template_workbook(
                    TemplateGenerationRequest(
                        config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                        output_path=path,
                        block_number="101",
                        fiscal_year="2026",
                        entry_rows=10,
                    )
                )

            set_sheet_value(node_a_path, "PST", "E4", 61)
            set_sheet_value(node_b_path, "PST", "E4", 62)

            with self.assertRaises(TemplateCompressionConflictError) as raised:
                compress_template_workbooks(
                    TemplateCompressionRequest(
                        config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                        workbook_paths=[node_a_path, node_b_path],
                        output_path=output_path,
                    )
                )

            message = str(raised.exception)
            self.assertIn("node_a.xlsx", message)
            self.assertIn("node_b.xlsx", message)
            self.assertIn("PST!E4 row 4", message)
            self.assertIn("61", message)
            self.assertIn("62", message)
            self.assertFalse(output_path.exists())


if __name__ == "__main__":
    unittest.main()
