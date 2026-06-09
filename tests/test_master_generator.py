from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = REPO_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from master_generator import (  # noqa: E402
    MasterGenerationRequest,
    generate_master_workbook,
    normalize_timed_value,
)
from template_generator import (  # noqa: E402
    TemplateGenerationRequest,
    generate_template_workbook,
    load_generation_inputs,
)


def set_sheet_value(path: Path, sheet_name: str, cell_ref: str, value: object) -> None:
    workbook = load_workbook(path)
    try:
        ws = workbook[sheet_name]
        ws[cell_ref] = value
        workbook.save(path)
    finally:
        workbook.close()


def rename_sheet(path: Path, old_name: str, new_name: str) -> None:
    workbook = load_workbook(path)
    try:
        workbook[old_name].title = new_name
        workbook.save(path)
    finally:
        workbook.close()


def set_metric_value(
    path: Path,
    metric_id: str,
    value: object,
    *,
    row_index: int = 4,
    occurrence: int = 1,
    sheet_name: str | None = None,
) -> None:
    workbook = load_workbook(path)
    try:
        sheets = [sheet_name] if sheet_name else workbook.sheetnames
        seen = 0
        for candidate_sheet_name in sheets:
            if candidate_sheet_name not in workbook.sheetnames:
                continue
            if candidate_sheet_name in {"META", "ROSTER", "LOOKUPS", "MASTER"}:
                continue
            ws = workbook[candidate_sheet_name]
            for col_idx in range(1, ws.max_column + 1):
                if ws.cell(row=3, column=col_idx).value == metric_id:
                    seen += 1
                    if seen == occurrence:
                        ws.cell(row=row_index, column=col_idx, value=value)
                        workbook.save(path)
                        return
        raise ValueError(
            "Metric {0} occurrence {1} not found".format(metric_id, occurrence)
        )
    finally:
        workbook.close()


def column_for_machine_header(ws: object, machine_header: str) -> int:
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=2, column=col_idx).value == machine_header:
            return col_idx
    raise ValueError("Machine header not found: {0}".format(machine_header))


class MasterGeneratorTest(unittest.TestCase):
    def test_generate_master_workbook_from_single_node_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            node_path = temp_root / "node_a.xlsx"
            master_path = temp_root / "master.xlsx"

            generate_template_workbook(
                TemplateGenerationRequest(
                    config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                    output_path=node_path,
                    block_number="101",
                    fiscal_year="2026",
                    entry_rows=10,
                )
            )

            set_sheet_value(node_path, "PST", "D1", "Push Ups Custom Label")
            rename_sheet(node_path, "PST", "PST Custom")
            set_sheet_value(node_path, "PST Custom", "D4", 61)
            set_sheet_value(node_path, "PST Custom", "E4", 62)
            set_sheet_value(node_path, "PST Custom", "F4", 13)
            set_sheet_value(node_path, "PST Custom", "G4", 10.30)
            set_sheet_value(node_path, "PST Custom", "H4", "12.00")

            generated_path = generate_master_workbook(
                MasterGenerationRequest(
                    config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                    workbook_paths=[node_path],
                    output_path=master_path,
                )
            )

            self.assertEqual(generated_path, master_path.resolve())
            workbook = load_workbook(generated_path, data_only=True)
            try:
                ws = workbook["MASTER"]
                loaded = load_generation_inputs(
                    TemplateGenerationRequest(
                        config_path=(REPO_ROOT / "config" / "config.yaml").resolve()
                    )
                )
                first_roster_row = loaded.roster_rows[0]
                self.assertIsNotNone(ws["A2"].value)
                self.assertEqual(ws["B2"].value, first_roster_row["last"])
                self.assertEqual(ws["C2"].value, first_roster_row["first"])
                self.assertEqual(ws["D2"].value, "101")
                self.assertEqual(ws["E2"].value, "node_a.xlsx")
                self.assertIsNone(ws["G2"].value)
                self.assertEqual(ws["H2"].value, 61)
                self.assertEqual(ws["I2"].value, 62)
                self.assertEqual(ws["J2"].value, 13)
                self.assertEqual(ws["K2"].value, 630)
                self.assertEqual(ws["L2"].value, 720)
            finally:
                workbook.close()

    def test_timed_values_support_mm_ss_entry_style(self) -> None:
        self.assertEqual(normalize_timed_value(10.30, entry_style="mm_ss"), 630)
        self.assertEqual(normalize_timed_value("10.30", entry_style="mm_ss"), 630)
        self.assertEqual(normalize_timed_value("10:30", entry_style="mm_ss"), 630)
        self.assertEqual(normalize_timed_value(12, entry_style="mm_ss"), 720)
        self.assertEqual(
            normalize_timed_value(600 / 86400, entry_style="mm_ss"),
            600,
        )
        with self.assertRaises(ValueError):
            normalize_timed_value("10.60", entry_style="mm_ss")

    def test_early_exit_rows_are_set_aside(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            node_path = temp_root / "node_a.xlsx"
            master_path = temp_root / "master.xlsx"

            generate_template_workbook(
                TemplateGenerationRequest(
                    config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                    output_path=node_path,
                    block_number="101",
                    fiscal_year="2026",
                    entry_rows=10,
                )
            )
            set_sheet_value(node_path, "PST", "D4", "MED PULL")
            set_sheet_value(node_path, "PST", "E4", 55)
            set_sheet_value(node_path, "PST", "D5", 45)

            generated_path = generate_master_workbook(
                MasterGenerationRequest(
                    config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                    workbook_paths=[node_path],
                    output_path=master_path,
                )
            )

            workbook = load_workbook(generated_path, data_only=True)
            try:
                loaded = load_generation_inputs(
                    TemplateGenerationRequest(
                        config_path=(REPO_ROOT / "config" / "config.yaml").resolve()
                    )
                )
                first_roster_row = loaded.roster_rows[0]
                second_roster_row = loaded.roster_rows[1]

                master_ws = workbook["MASTER"]
                self.assertEqual(master_ws.max_row, len(loaded.roster_rows))
                self.assertEqual(master_ws["A2"].value, second_roster_row["uid"])
                self.assertEqual(master_ws["H2"].value, 45)

                early_exit_ws = workbook["EARLY_EXITS"]
                self.assertEqual(early_exit_ws.max_row, 2)
                self.assertEqual(early_exit_ws["A2"].value, first_roster_row["uid"])
                self.assertEqual(early_exit_ws["F2"].value, "MED PULL")
                self.assertEqual(early_exit_ws["G2"].value, "node_a.xlsx")
                self.assertEqual(early_exit_ws["H2"].value, "PST")
            finally:
                workbook.close()

    def test_dynamic_master_flattens_metrics_without_master_config(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            node_a_path = temp_root / "node_a.xlsx"
            node_b_path = temp_root / "node_b.xlsx"
            master_path = temp_root / "dynamic_master.xlsx"

            for path in (node_a_path, node_b_path):
                generate_template_workbook(
                    TemplateGenerationRequest(
                        config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                        output_path=path,
                        block_number="101",
                        fiscal_year="2026",
                        entry_rows=10,
                    )
                )

            set_metric_value(node_a_path, "m_push_ups", 10)
            set_metric_value(node_b_path, "m_push_ups", 20)
            set_metric_value(node_a_path, "m_ocourse_pass_fail", "Pass")
            set_metric_value(node_b_path, "m_ocourse_pass_fail", "Fail")
            set_metric_value(
                node_a_path,
                "m_grit_pt_physicality",
                4,
                sheet_name="Grit PT #1",
            )
            set_metric_value(node_a_path, "m_push_ups", 50, row_index=5)
            set_metric_value(node_b_path, "m_push_ups", "DOR", row_index=5)

            generated_path = generate_master_workbook(
                MasterGenerationRequest(
                    config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                    workbook_paths=[node_a_path, node_b_path],
                    output_path=master_path,
                    dynamic=True,
                )
            )

            workbook = load_workbook(generated_path, data_only=True)
            try:
                loaded = load_generation_inputs(
                    TemplateGenerationRequest(
                        config_path=(REPO_ROOT / "config" / "config.yaml").resolve()
                    )
                )
                first_roster_row = loaded.roster_rows[0]
                second_roster_row = loaded.roster_rows[1]
                ws = workbook["MASTER"]
                self.assertTrue(ws.row_dimensions[2].hidden)

                machine_headers = [
                    ws.cell(row=2, column=col_idx).value
                    for col_idx in range(1, ws.max_column + 1)
                ]
                metric_headers = machine_headers[6:]
                self.assertEqual(len(metric_headers), len(set(metric_headers)))

                ids = [
                    ws.cell(row=row_idx, column=1).value
                    for row_idx in range(3, ws.max_row + 1)
                ]
                self.assertIn(first_roster_row["uid"], ids)
                self.assertNotIn(second_roster_row["uid"], ids)
                first_row_idx = ids.index(first_roster_row["uid"]) + 3

                push_ups_col = column_for_machine_header(ws, "m_push_ups")
                pass_fail_col = column_for_machine_header(ws, "m_ocourse_pass_fail")
                subjective_col = column_for_machine_header(
                    ws,
                    "m_grit_pt_physicality",
                )

                self.assertEqual(
                    ws.cell(row=first_row_idx, column=push_ups_col).value,
                    15,
                )
                self.assertEqual(
                    ws.cell(row=first_row_idx, column=pass_fail_col).value,
                    "Pass; Fail",
                )
                self.assertAlmostEqual(
                    ws.cell(row=first_row_idx, column=subjective_col).value,
                    14 / 6,
                )
                self.assertGreater(len(ws.conditional_formatting), 0)

                early_exit_ws = workbook["EARLY_EXITS"]
                self.assertEqual(early_exit_ws["A2"].value, second_roster_row["uid"])
                self.assertEqual(early_exit_ws["F2"].value, "DOR")
            finally:
                workbook.close()

    def test_aggregate_duplicate_metric_values_across_workbooks(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            node_a_path = temp_root / "node_a.xlsx"
            node_b_path = temp_root / "node_b.xlsx"
            master_path = temp_root / "master.xlsx"

            for path in (node_a_path, node_b_path):
                generate_template_workbook(
                    TemplateGenerationRequest(
                        config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                        output_path=path,
                        block_number="101",
                        fiscal_year="2026",
                        entry_rows=10,
                    )
                )

            set_sheet_value(node_a_path, "PST", "D4", 45)
            set_sheet_value(node_b_path, "PST", "E4", 55)
            set_sheet_value(node_a_path, "Log PT", "D4", 3)
            set_sheet_value(node_b_path, "Log Carry around O Course", "D4", 5)

            generated_path = generate_master_workbook(
                MasterGenerationRequest(
                    config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                    workbook_paths=[node_a_path, node_b_path],
                    output_path=master_path,
                )
            )

            workbook = load_workbook(generated_path, data_only=True)
            try:
                ws = workbook["MASTER"]
                loaded = load_generation_inputs(
                    TemplateGenerationRequest(
                        config_path=(REPO_ROOT / "config" / "config.yaml").resolve()
                    )
                )
                self.assertEqual(ws.max_row, len(loaded.roster_rows) + 1)
                self.assertEqual(ws["H2"].value, 45)
                self.assertEqual(ws["I2"].value, 55)
                self.assertEqual(ws["T2"].value, 4)
                self.assertEqual(ws["E2"].value, "node_a.xlsx; node_b.xlsx")
            finally:
                workbook.close()

    def test_generate_master_workbook_from_dropbox_is_cumulative(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            dropbox_path = temp_root / "dropbox"
            dropbox_path.mkdir()
            master_path = temp_root / "master.xlsx"
            node_a_path = dropbox_path / "node_a.xlsx"

            generate_template_workbook(
                TemplateGenerationRequest(
                    config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                    output_path=node_a_path,
                    block_number="101",
                    fiscal_year="2026",
                    entry_rows=10,
                )
            )
            set_sheet_value(node_a_path, "PST", "D4", 45)

            generate_master_workbook(
                MasterGenerationRequest(
                    config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                    dropbox_path=dropbox_path,
                    output_path=master_path,
                )
            )

            processed_path = dropbox_path / "processed"
            self.assertFalse(node_a_path.exists())
            self.assertTrue((processed_path / "node_a.xlsx").exists())

            node_b_path = dropbox_path / "node_b.xlsx"
            generate_template_workbook(
                TemplateGenerationRequest(
                    config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                    output_path=node_b_path,
                    block_number="101",
                    fiscal_year="2026",
                    entry_rows=10,
                )
            )
            set_sheet_value(node_b_path, "PST", "E4", 55)

            generate_master_workbook(
                MasterGenerationRequest(
                    config_path=(REPO_ROOT / "config" / "config.yaml").resolve(),
                    dropbox_path=dropbox_path,
                    output_path=master_path,
                )
            )

            self.assertFalse(node_b_path.exists())
            self.assertTrue((processed_path / "node_b.xlsx").exists())
            workbook = load_workbook(master_path, data_only=True)
            try:
                ws = workbook["MASTER"]
                self.assertEqual(ws["E2"].value, "node_a.xlsx; node_b.xlsx")
                self.assertEqual(ws["H2"].value, 45)
                self.assertEqual(ws["I2"].value, 55)
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
