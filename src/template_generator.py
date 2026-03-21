from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from datetime import datetime, timezone
import hashlib
from pathlib import Path
import re
from typing import Any, Dict, List, Mapping, Optional, Set

import yaml
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)
INVALID_FILL = PatternFill(fill_type="solid", fgColor="F8CBAD")
INVALID_FONT = Font(color="9C0006")
DEFAULT_ENTRY_ROWS = 200
DEFAULT_CONFIG_PATH = "config/config.yaml"


@dataclass(frozen=True)
class TemplateGenerationRequest:
    config_path: Path
    roster_path: Optional[Path] = None
    output_path: Optional[Path] = None
    block_number: str = "TBD"
    fiscal_year: str = "TBD"
    entry_rows: int = DEFAULT_ENTRY_ROWS


@dataclass(frozen=True)
class LoadedTemplateConfig:
    config_path: Path
    metrics_path: Path
    evolutions_path: Path
    roster_path: Optional[Path]
    config_doc: Dict[str, Any]
    metrics_doc: Dict[str, Any]
    evolutions_doc: Dict[str, Any]
    roster_rows: List[Dict[str, str]]
    metrics_by_id: Dict[str, Dict[str, Any]]


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate an Excel node workbook from canonical YAML configs."
    )
    parser.add_argument(
        "--config",
        default=DEFAULT_CONFIG_PATH,
        help="Path to canonical config file.",
    )
    parser.add_argument(
        "--roster",
        default=None,
        help="Optional CSV with either uid,first,last or last,first,dob columns.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output xlsx path (default: workbooks/NodeTemplate_v{version}.xlsx).",
    )
    parser.add_argument(
        "--block-number",
        default="TBD",
        help="Block number to stamp in META/evolution sheets.",
    )
    parser.add_argument(
        "--fiscal-year",
        default="TBD",
        help="Fiscal year to stamp in META/evolution sheets.",
    )
    parser.add_argument(
        "--entry-rows",
        type=int,
        default=DEFAULT_ENTRY_ROWS,
        help="How many editable candidate rows to pre-unlock in each evolution sheet.",
    )
    return parser.parse_args(argv)


def request_from_namespace(args: argparse.Namespace) -> TemplateGenerationRequest:
    return TemplateGenerationRequest(
        config_path=Path(args.config).resolve(),
        roster_path=Path(args.roster).resolve() if args.roster else None,
        output_path=Path(args.output).resolve() if args.output else None,
        block_number=args.block_number,
        fiscal_year=args.fiscal_year,
        entry_rows=args.entry_rows,
    )


def load_yaml(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        data = yaml.safe_load(handle)
    if not isinstance(data, dict):
        raise ValueError("Expected mapping YAML at {0}".format(path))
    return data


def resolve_path(config_path: Path, candidate: str) -> Path:
    path_candidate = Path(candidate)
    if path_candidate.is_absolute():
        return path_candidate
    return (config_path.parent.parent / path_candidate).resolve()


def sanitize_defined_name(name: str) -> str:
    sanitized = re.sub(r"[^A-Za-z0-9_]", "_", name)
    if sanitized and sanitized[0].isdigit():
        sanitized = "_{0}".format(sanitized)
    return sanitized or "domain_values"


def normalize_roster_header(header: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "", header.strip().lower())
    aliases = {
        "uid": "uid",
        "first": "first",
        "firstname": "first",
        "last": "last",
        "lastname": "last",
        "dob": "dob",
        "dateofbirth": "dob",
    }
    return aliases.get(normalized, normalized)


def parse_roster_dob(raw_value: str, roster_path: Path) -> str:
    stripped = raw_value.strip()
    if not stripped:
        raise ValueError("Roster row is missing dob: {0}".format(roster_path))

    supported_formats = ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y")
    for date_format in supported_formats:
        try:
            return datetime.strptime(stripped, date_format).date().isoformat()
        except ValueError:
            continue
    raise ValueError(
        "Unsupported dob format '{0}' in {1}; use YYYY-MM-DD or MM/DD/YYYY".format(
            raw_value, roster_path
        )
    )


def build_candidate_uid(first: str, last: str, dob: str) -> str:
    normalized = "|".join(part.strip().casefold() for part in (last, first, dob))
    return hashlib.blake2b(normalized.encode("utf-8"), digest_size=12).hexdigest()


def load_roster(roster_path: Optional[Path]) -> List[Dict[str, str]]:
    if roster_path is None:
        return []

    with roster_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError("Roster file has no headers: {0}".format(roster_path))

        rows = []  # type: List[Dict[str, str]]
        seen_uids = set()  # type: Set[str]
        canonical_headers = {
            header: normalize_roster_header(header)
            for header in reader.fieldnames
            if header
        }
        header_values = set(canonical_headers.values())

        if {"uid", "first", "last"} <= header_values:
            roster_mode = "uid"
        elif {"dob", "first", "last"} <= header_values:
            roster_mode = "dob"
        else:
            raise ValueError(
                "Roster must include either uid,first,last or last,first,dob columns"
            )

        for row_number, row in enumerate(reader, start=2):
            cleaned = {}
            for key, value in row.items():
                if key and canonical_headers.get(key):
                    cleaned[canonical_headers[key]] = (value or "").strip()

            first = cleaned.get("first", "")
            last = cleaned.get("last", "")
            if not first or not last:
                raise ValueError(
                    "Roster row {0} in {1} is missing first/last".format(
                        row_number, roster_path
                    )
                )

            if roster_mode == "uid":
                uid = cleaned.get("uid", "")
                if not uid:
                    raise ValueError(
                        "Roster row {0} in {1} is missing uid".format(
                            row_number, roster_path
                        )
                    )
                dob = cleaned.get("dob", "")
            else:
                dob = parse_roster_dob(cleaned.get("dob", ""), roster_path)
                uid = build_candidate_uid(first=first, last=last, dob=dob)

            if uid in seen_uids:
                raise ValueError(
                    "Duplicate roster uid '{0}' in {1}".format(uid, roster_path)
                )
            seen_uids.add(uid)
            rows.append({"uid": uid, "first": first, "last": last, "dob": dob})
    return rows


def build_metric_index(metrics_doc: Mapping[str, Any]) -> Dict[str, Dict[str, Any]]:
    metrics = metrics_doc.get("metrics")
    if not isinstance(metrics, list):
        raise ValueError("metrics.yaml must contain a list under 'metrics'")

    index = {}  # type: Dict[str, Dict[str, Any]]
    for metric in metrics:
        if not isinstance(metric, dict):
            raise ValueError("Each metric entry must be a mapping")
        metric_id = metric.get("metric_id")
        if not metric_id:
            raise ValueError("Metric missing metric_id")
        if metric_id in index:
            raise ValueError("Duplicate metric_id: {0}".format(metric_id))
        index[str(metric_id)] = metric
    return index


def validate_metric_definitions(
    config_doc: Mapping[str, Any], metrics_by_id: Mapping[str, Mapping[str, Any]]
) -> None:
    validation = config_doc.get("validation", {})
    if validation is None:
        validation = {}
    if not isinstance(validation, dict):
        raise ValueError("config.yaml 'validation' must be a mapping")

    required_metric_fields = validation.get("required_metric_fields", [])
    if not isinstance(required_metric_fields, list):
        raise ValueError("config.yaml 'validation.required_metric_fields' must be a list")

    allowed_metric_types = validation.get("allowed_metric_types", [])
    if not isinstance(allowed_metric_types, list):
        raise ValueError("config.yaml 'validation.allowed_metric_types' must be a list")

    allowed_input_kinds = validation.get("allowed_input_kinds", [])
    if not isinstance(allowed_input_kinds, list):
        raise ValueError("config.yaml 'validation.allowed_input_kinds' must be a list")

    allowed_metric_types_set = set(str(value) for value in allowed_metric_types)
    allowed_input_kinds_set = set(str(value) for value in allowed_input_kinds)

    for metric_id, metric in metrics_by_id.items():
        for field_name in required_metric_fields:
            if field_name not in metric or metric.get(field_name) in ("", None):
                raise ValueError(
                    "Metric '{0}' is missing required field '{1}'".format(
                        metric_id, field_name
                    )
                )

        metric_type = metric.get("type")
        if allowed_metric_types_set and metric_type not in allowed_metric_types_set:
            raise ValueError(
                "Metric '{0}' uses unsupported type '{1}'".format(metric_id, metric_type)
            )

        input_kind = metric.get("input_kind")
        if allowed_input_kinds_set and input_kind not in allowed_input_kinds_set:
            raise ValueError(
                "Metric '{0}' uses unsupported input_kind '{1}'".format(
                    metric_id, input_kind
                )
            )


def validate_contract(
    config_doc: Mapping[str, Any],
    metrics_by_id: Mapping[str, Mapping[str, Any]],
    evolutions_doc: Mapping[str, Any],
) -> None:
    domains = config_doc.get("domains", {})
    if not isinstance(domains, dict):
        raise ValueError("config.yaml 'domains' must be a mapping")

    for metric_id, metric in metrics_by_id.items():
        domain_ref = metric.get("domain_ref")
        if domain_ref and domain_ref not in domains:
            raise ValueError(
                "Metric '{0}' references unknown domain '{1}'".format(
                    metric_id, domain_ref
                )
            )

        derived_from = metric.get("derived_from", [])
        if derived_from:
            for dependency in derived_from:
                if dependency not in metrics_by_id:
                    raise ValueError(
                        "Metric '{0}' references unknown derived input '{1}'".format(
                            metric_id, dependency
                        )
                    )

    evolutions = evolutions_doc.get("evolutions")
    if not isinstance(evolutions, list) or not evolutions:
        raise ValueError("evolutions.yaml must contain a non-empty 'evolutions' list")

    seen_evolutions = set()  # type: Set[str]
    for evolution in evolutions:
        evolution_id = evolution.get("evolution_id")
        if not evolution_id:
            raise ValueError("Evolution missing evolution_id")
        if evolution_id in seen_evolutions:
            raise ValueError("Duplicate evolution_id: {0}".format(evolution_id))
        seen_evolutions.add(str(evolution_id))

        metric_ids = evolution.get("metric_ids", [])
        if not metric_ids:
            raise ValueError(
                "Evolution '{0}' has no metric_ids".format(evolution_id)
            )
        for metric_id in metric_ids:
            if metric_id not in metrics_by_id:
                raise ValueError(
                    "Evolution '{0}' references unknown metric_id '{1}'".format(
                        evolution_id, metric_id
                    )
                )


def validate_generation_inputs(loaded: LoadedTemplateConfig) -> None:
    files = loaded.config_doc.get("files", {})
    if files is None:
        files = {}
    if not isinstance(files, dict):
        raise ValueError("config.yaml 'files' must be a mapping")

    sheet_contract = loaded.config_doc.get("sheet_contract", {})
    if sheet_contract is None:
        sheet_contract = {}
    if not isinstance(sheet_contract, dict):
        raise ValueError("config.yaml 'sheet_contract' must be a mapping")

    validate_metric_definitions(
        config_doc=loaded.config_doc,
        metrics_by_id=loaded.metrics_by_id,
    )
    validate_contract(
        config_doc=loaded.config_doc,
        metrics_by_id=loaded.metrics_by_id,
        evolutions_doc=loaded.evolutions_doc,
    )


def load_generation_inputs(
    request: TemplateGenerationRequest,
) -> LoadedTemplateConfig:
    if request.entry_rows < 1:
        raise ValueError("entry_rows must be at least 1")

    config_path = request.config_path.resolve()
    config_doc = load_yaml(config_path)

    files = config_doc.get("files", {})
    if files is None:
        files = {}
    if not isinstance(files, dict):
        raise ValueError("config.yaml 'files' must be a mapping")

    metrics_path = resolve_path(config_path, files.get("metrics", "config/metrics.yaml"))
    evolutions_path = resolve_path(
        config_path, files.get("evolutions", "config/evolutions.yaml")
    )
    roster_path = request.roster_path
    if roster_path is None:
        roster_candidate = files.get("roster")
        if roster_candidate:
            roster_path = resolve_path(config_path, roster_candidate)

    metrics_doc = load_yaml(metrics_path)
    evolutions_doc = load_yaml(evolutions_path)
    roster_rows = load_roster(roster_path)
    metrics_by_id = build_metric_index(metrics_doc)

    loaded = LoadedTemplateConfig(
        config_path=config_path,
        metrics_path=metrics_path,
        evolutions_path=evolutions_path,
        roster_path=roster_path,
        config_doc=config_doc,
        metrics_doc=metrics_doc,
        evolutions_doc=evolutions_doc,
        roster_rows=roster_rows,
        metrics_by_id=metrics_by_id,
    )
    validate_generation_inputs(loaded)
    return loaded


def style_header_row(ws: Any, headers: List[str], row_index: int) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_index, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.protection = Protection(locked=True)


def create_meta_sheet(
    wb: Workbook,
    config_doc: Mapping[str, Any],
    block_number: str,
    fiscal_year: str,
) -> None:
    ws = wb.create_sheet("META")
    rows = [
        ("registry_name", config_doc.get("registry_name", "")),
        ("config_version", config_doc.get("version", "")),
        ("generated_on_utc", datetime.now(timezone.utc).isoformat()),
        ("block_number", block_number),
        ("fiscal_year", fiscal_year),
    ]
    for row_index, (key, value) in enumerate(rows, start=1):
        ws.cell(row=row_index, column=1, value=key).font = HEADER_FONT
        ws.cell(row=row_index, column=2, value=value)
        ws.cell(row=row_index, column=1).protection = Protection(locked=True)
        ws.cell(row=row_index, column=2).protection = Protection(locked=True)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 36
    ws.protection.sheet = True
    ws.protection.enable()


def create_roster_sheet(wb: Workbook, roster_rows: List[Dict[str, str]]) -> None:
    ws = wb.create_sheet("ROSTER")
    headers = ["UID", "First", "Last"]
    style_header_row(ws, headers, 1)

    for row_index, row in enumerate(roster_rows, start=2):
        ws.cell(row=row_index, column=1, value=row.get("uid"))
        ws.cell(row=row_index, column=2, value=row.get("first"))
        ws.cell(row=row_index, column=3, value=row.get("last"))

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    editable_row_end = max(2, len(roster_rows) + DEFAULT_ENTRY_ROWS)
    for row_index in range(2, editable_row_end + 1):
        for col_idx in range(1, 4):
            ws.cell(row=row_index, column=col_idx).protection = Protection(locked=False)
    ws.protection.sheet = True
    ws.protection.enable()


def create_lookups_sheet(
    wb: Workbook, domains: Mapping[str, List[Any]]
) -> Dict[str, str]:
    ws = wb.create_sheet("LOOKUPS")
    named_ranges = {}  # type: Dict[str, str]

    for col_idx, (domain, values) in enumerate(domains.items(), start=1):
        ws.cell(row=1, column=col_idx, value=domain).font = HEADER_FONT
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(14, len(domain) + 4)

        for row_idx, value in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

        if values:
            defined_name = "domain_{0}".format(sanitize_defined_name(domain))
            ref = "'LOOKUPS'!${0}$2:${0}${1}".format(col_letter, len(values) + 1)
            wb.defined_names.add(DefinedName(name=defined_name, attr_text=ref))
            named_ranges[domain] = defined_name

    ws.sheet_state = "hidden"
    ws.protection.sheet = True
    ws.protection.enable()
    return named_ranges


def add_invalid_highlight_rule(ws: Any, formula: str, cell_range: str) -> None:
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[formula], stopIfTrue=False, fill=INVALID_FILL, font=INVALID_FONT),
    )


def _try_float(value: Any) -> Optional[float]:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            return float(stripped)
        except ValueError:
            return None
    return None


def add_min_pass_highlight_rule(
    ws: Any,
    metric: Mapping[str, Any],
    first_cell: str,
    cell_range: str,
) -> None:
    min_pass = metric.get("min_pass")
    if min_pass is None:
        return

    metric_type = metric.get("type")
    threshold = _try_float(min_pass)
    if threshold is None:
        return

    if metric_type == "timed":
        excel_threshold = threshold / 86400
        formula = "AND(NOT(ISBLANK({0})),ISNUMBER({0}),{0}>{1})".format(
            first_cell, excel_threshold
        )
    else:
        formula = "AND(NOT(ISBLANK({0})),ISNUMBER({0}),{0}<{1})".format(
            first_cell, threshold
        )

    add_invalid_highlight_rule(ws=ws, formula=formula, cell_range=cell_range)


def add_data_validations(
    ws: Any,
    metric: Mapping[str, Any],
    col_idx: int,
    row_start: int,
    row_end: int,
    domain_named_ranges: Mapping[str, str],
) -> None:
    col_letter = get_column_letter(col_idx)
    cell_range = "{0}{1}:{0}{2}".format(col_letter, row_start, row_end)
    first_cell = "{0}{1}".format(col_letter, row_start)
    metric_type = metric.get("type")
    domain_ref = metric.get("domain_ref")

    if domain_ref and domain_ref in domain_named_ranges:
        dv = DataValidation(
            type="list",
            formula1="={0}".format(domain_named_ranges[domain_ref]),
            allow_blank=True,
        )
        dv.error = "Value must match one of the configured options."
        dv.errorTitle = "Invalid Dropdown Value"
        ws.add_data_validation(dv)
        dv.add(cell_range)
        add_invalid_highlight_rule(
            ws=ws,
            formula=(
                "AND(NOT(ISBLANK({0})),COUNTIF({1},{0})=0)".format(
                    first_cell, domain_named_ranges[domain_ref]
                )
            ),
            cell_range=cell_range,
        )

    add_min_pass_highlight_rule(
        ws=ws,
        metric=metric,
        first_cell=first_cell,
        cell_range=cell_range,
    )

    if metric_type == "timed":
        for row_idx in range(row_start, row_end + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = "[mm]:ss"


def evolution_sheet_name(evolution: Mapping[str, Any]) -> str:
    for key in ("sheet_name", "display_name", "evolution_id"):
        value = evolution.get(key)
        if value:
            return str(value)[:31]
    return "Evolution"


def create_evolution_sheets(
    wb: Workbook,
    evolutions_doc: Mapping[str, Any],
    metrics_by_id: Mapping[str, Mapping[str, Any]],
    roster_rows: List[Dict[str, str]],
    sheet_contract: Mapping[str, Any],
    block_number: str,
    fiscal_year: str,
    entry_rows: int,
    domain_named_ranges: Mapping[str, str],
) -> None:
    configured_roster_fields = sheet_contract.get(
        "locked_left_columns", ["uid", "first", "last"]
    )
    roster_fields = ["uid", "first", "last"] + [
        field
        for field in configured_roster_fields
        if field not in {"uid", "first", "last"}
    ]
    header_row = int(sheet_contract.get("header_row", 1))
    first_candidate_row = int(sheet_contract.get("first_candidate_row", 2))
    freeze_panes = sheet_contract.get("freeze_panes", "D2")

    edit_row_count = max(len(roster_rows), entry_rows)
    edit_row_end = first_candidate_row + edit_row_count - 1

    for evolution in evolutions_doc["evolutions"]:
        ws = wb.create_sheet(evolution_sheet_name(evolution))

        metric_ids = evolution.get("metric_ids", [])
        roster_headers = {"uid": "UID", "first": "First", "last": "Last"}
        headers = [roster_headers.get(field, field.capitalize()) for field in roster_fields]
        headers.extend(
            metrics_by_id[metric_id].get("display_name", metric_id)
            for metric_id in metric_ids
        )
        style_header_row(ws, headers, header_row)
        ws.freeze_panes = freeze_panes

        for row_idx, roster in enumerate(roster_rows, start=first_candidate_row):
            for col_idx, field in enumerate(roster_fields, start=1):
                ws.cell(row=row_idx, column=col_idx, value=roster.get(field, ""))

        meta_start_col = len(headers) + 2
        ws.cell(row=1, column=meta_start_col, value="evolution_id").font = HEADER_FONT
        ws.cell(row=1, column=meta_start_col + 1, value=evolution["evolution_id"])
        ws.cell(row=2, column=meta_start_col, value="block_number").font = HEADER_FONT
        ws.cell(row=2, column=meta_start_col + 1, value=block_number)
        ws.cell(row=3, column=meta_start_col, value="fiscal_year").font = HEADER_FONT
        ws.cell(row=3, column=meta_start_col + 1, value=fiscal_year)
        ws.cell(row=1, column=meta_start_col).protection = Protection(locked=True)
        ws.cell(row=1, column=meta_start_col + 1).protection = Protection(locked=True)
        ws.cell(row=2, column=meta_start_col).protection = Protection(locked=True)
        ws.cell(row=2, column=meta_start_col + 1).protection = Protection(locked=True)
        ws.cell(row=3, column=meta_start_col).protection = Protection(locked=True)
        ws.cell(row=3, column=meta_start_col + 1).protection = Protection(locked=True)

        metric_start_col = len(roster_fields) + 1
        for offset, metric_id in enumerate(metric_ids):
            col_idx = metric_start_col + offset
            metric = metrics_by_id[metric_id]
            add_data_validations(
                ws=ws,
                metric=metric,
                col_idx=col_idx,
                row_start=first_candidate_row,
                row_end=edit_row_end,
                domain_named_ranges=domain_named_ranges,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = 24

        for col_idx in range(1, len(roster_fields) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

        ws.protection.sheet = True
        ws.protection.enable()
        for row_idx in range(first_candidate_row, edit_row_end + 1):
            for col_idx in range(1, len(roster_fields) + 1):
                ws.cell(row=row_idx, column=col_idx).protection = Protection(locked=True)
            for col_idx in range(metric_start_col, metric_start_col + len(metric_ids)):
                ws.cell(row=row_idx, column=col_idx).protection = Protection(locked=False)


def default_output_path(loaded: LoadedTemplateConfig) -> Path:
    return (
        loaded.config_path.parent.parent
        / "workbooks"
        / "NodeTemplate_v{0}.xlsx".format(loaded.config_doc.get("version", "unknown"))
    ).resolve()


def build_workbook(
    loaded: LoadedTemplateConfig, request: TemplateGenerationRequest
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    create_meta_sheet(
        wb=wb,
        config_doc=loaded.config_doc,
        block_number=request.block_number,
        fiscal_year=request.fiscal_year,
    )
    create_roster_sheet(wb=wb, roster_rows=loaded.roster_rows)
    domain_named_ranges = create_lookups_sheet(
        wb=wb, domains=loaded.config_doc.get("domains", {})
    )
    create_evolution_sheets(
        wb=wb,
        evolutions_doc=loaded.evolutions_doc,
        metrics_by_id=loaded.metrics_by_id,
        roster_rows=loaded.roster_rows,
        sheet_contract=loaded.config_doc.get("sheet_contract", {}),
        block_number=request.block_number,
        fiscal_year=request.fiscal_year,
        entry_rows=max(1, request.entry_rows),
        domain_named_ranges=domain_named_ranges,
    )
    return wb


def generate_template_workbook(request: TemplateGenerationRequest) -> Path:
    loaded = load_generation_inputs(request)
    workbook = build_workbook(loaded, request)
    output_path = request.output_path.resolve() if request.output_path else default_output_path(loaded)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def generate_workbook(request: TemplateGenerationRequest) -> Path:
    return generate_template_workbook(request)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    output_path = generate_template_workbook(request_from_namespace(args))
    print("Workbook generated: {0}".format(output_path))
    return 0

