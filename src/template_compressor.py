from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from master_generator import (
    dedupe_paths,
    detect_machine_header_row,
    discover_workbooks,
    is_excel_workbook,
    metric_columns_for_sheet,
    read_candidate_cell,
    sheet_row_range,
)
from template_generator import (
    DEFAULT_CONFIG_PATH,
    TemplateGenerationRequest,
    configured_locked_left_columns,
    load_generation_inputs,
)


SYSTEM_SHEETS = {"META", "ROSTER", "LOOKUPS", "MASTER", "EARLY_EXITS"}


@dataclass(frozen=True)
class TemplateCompressionRequest:
    config_path: Path
    workbook_paths: Sequence[Path] = ()
    output_path: Optional[Path] = None
    dropbox_path: Optional[Path] = None


@dataclass(frozen=True)
class ScoreCellConflict:
    sheet_name: str
    row_index: int
    column_index: int
    existing_workbook: str
    incoming_workbook: str
    existing_value: Any
    incoming_value: Any

    @property
    def cell_ref(self) -> str:
        return "{0}{1}".format(get_column_letter(self.column_index), self.row_index)


class TemplateCompressionConflictError(ValueError):
    def __init__(self, conflicts: Sequence[ScoreCellConflict]) -> None:
        self.conflicts = list(conflicts)
        super().__init__(format_conflict_message(self.conflicts))


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compress multiple scored node templates into one node workbook."
    )
    parser.add_argument(
        "--config",
        default=DEFAULT_CONFIG_PATH,
        help="Path to canonical config file.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help=(
            "Output xlsx path "
            "(default: workbooks/CompressedNodeTemplate_v{version}.xlsx)."
        ),
    )
    parser.add_argument(
        "--dropbox",
        default=None,
        help="Directory containing node workbooks to compress.",
    )
    parser.add_argument(
        "workbooks",
        nargs="*",
        help="Optional explicit node workbook xlsx files to compress.",
    )
    return parser.parse_args(argv)


def request_from_namespace(args: argparse.Namespace) -> TemplateCompressionRequest:
    return TemplateCompressionRequest(
        config_path=Path(args.config).resolve(),
        workbook_paths=[Path(value).resolve() for value in args.workbooks],
        output_path=Path(args.output).resolve() if args.output else None,
        dropbox_path=Path(args.dropbox).resolve() if args.dropbox else None,
    )


def default_output_path(request: TemplateCompressionRequest) -> Path:
    loaded = load_generation_inputs(
        TemplateGenerationRequest(config_path=request.config_path)
    )
    return (
        loaded.config_path.parent.parent
        / "workbooks"
        / "CompressedNodeTemplate_v{0}.xlsx".format(
            loaded.config_doc.get("version", "unknown")
        )
    ).resolve()


def resolve_compression_workbooks(request: TemplateCompressionRequest) -> List[Path]:
    paths = list(Path(path).resolve() for path in request.workbook_paths)
    if request.dropbox_path is not None:
        dropbox_path = request.dropbox_path.resolve()
        if not dropbox_path.is_dir():
            raise ValueError("Dropbox path is not a directory: {0}".format(dropbox_path))
        paths.extend(discover_workbooks(dropbox_path))

    resolved = dedupe_paths(paths)
    for workbook_path in resolved:
        if not is_excel_workbook(workbook_path):
            raise ValueError(
                "Node workbook path is not a readable .xlsx file: {0}".format(
                    workbook_path
                )
            )
    if not resolved:
        raise ValueError("Provide at least one workbook path or --dropbox")
    return resolved


def is_blank_value(value: Any) -> bool:
    return value is None or value == ""


def values_match(left: Any, right: Any) -> bool:
    return left == right


def all_configured_roster_fields_from_template(loaded: Any) -> List[str]:
    sheet_contract = loaded.config_doc.get("sheet_contract", {})
    roster_fields = configured_locked_left_columns(sheet_contract)
    for evolution in loaded.evolutions_doc.get("evolutions", []):
        if not isinstance(evolution, dict):
            continue
        for field_name in configured_locked_left_columns(sheet_contract, evolution):
            if field_name not in roster_fields:
                roster_fields.append(field_name)
    return roster_fields


def score_sheet_names(
    *,
    workbook: Any,
    configured_metric_id_row: int,
    metrics_by_id: Mapping[str, Mapping[str, Any]],
    roster_fields: Sequence[str],
) -> List[str]:
    sheet_names = []
    for sheet_name in workbook.sheetnames:
        if sheet_name in SYSTEM_SHEETS:
            continue
        ws = workbook[sheet_name]
        metric_id_row = detect_machine_header_row(
            ws=ws,
            configured_metric_id_row=configured_metric_id_row,
            metrics_by_id=metrics_by_id,
            roster_fields=roster_fields,
        )
        if metric_columns_for_sheet(
            ws=ws,
            metric_id_row=metric_id_row,
            metrics_by_id=metrics_by_id,
        ):
            sheet_names.append(sheet_name)
    return sheet_names


def validate_matching_score_sheets(
    *,
    base_workbook: Any,
    source_workbook: Any,
    source_path: Path,
    configured_metric_id_row: int,
    metrics_by_id: Mapping[str, Mapping[str, Any]],
    roster_fields: Sequence[str],
) -> None:
    base_sheets = score_sheet_names(
        workbook=base_workbook,
        configured_metric_id_row=configured_metric_id_row,
        metrics_by_id=metrics_by_id,
        roster_fields=roster_fields,
    )
    source_sheets = score_sheet_names(
        workbook=source_workbook,
        configured_metric_id_row=configured_metric_id_row,
        metrics_by_id=metrics_by_id,
        roster_fields=roster_fields,
    )
    if set(base_sheets) != set(source_sheets):
        raise ValueError(
            "Workbook {0} does not match base template sheets. "
            "Expected {1}, found {2}".format(
                source_path.name,
                base_sheets,
                source_sheets,
            )
        )


def metric_columns_for_matching_sheet(
    *,
    ws: Any,
    configured_metric_id_row: int,
    metrics_by_id: Mapping[str, Mapping[str, Any]],
    roster_fields: Sequence[str],
) -> Tuple[int, Dict[int, str]]:
    metric_id_row = detect_machine_header_row(
        ws=ws,
        configured_metric_id_row=configured_metric_id_row,
        metrics_by_id=metrics_by_id,
        roster_fields=roster_fields,
    )
    return (
        metric_id_row,
        metric_columns_for_sheet(
            ws=ws,
            metric_id_row=metric_id_row,
            metrics_by_id=metrics_by_id,
        ),
    )


def validate_matching_metric_columns(
    *,
    sheet_name: str,
    base_columns: Mapping[int, str],
    source_columns: Mapping[int, str],
    source_path: Path,
) -> None:
    if dict(base_columns) != dict(source_columns):
        raise ValueError(
            "Workbook {0} sheet {1} does not match base metric columns. "
            "Expected {2}, found {3}".format(
                source_path.name,
                sheet_name,
                dict(base_columns),
                dict(source_columns),
            )
        )


def initialize_source_map(
    *,
    base_workbook: Any,
    base_path: Path,
    score_sheets: Sequence[str],
    configured_metric_id_row: int,
    configured_first_candidate_row: int,
    metrics_by_id: Mapping[str, Mapping[str, Any]],
    roster_fields: Sequence[str],
) -> Dict[Tuple[str, int, int], str]:
    source_map = {}
    for sheet_name in score_sheets:
        ws = base_workbook[sheet_name]
        metric_id_row, metric_columns = metric_columns_for_matching_sheet(
            ws=ws,
            configured_metric_id_row=configured_metric_id_row,
            metrics_by_id=metrics_by_id,
            roster_fields=roster_fields,
        )
        first_candidate_row = max(configured_first_candidate_row, metric_id_row + 1)
        for row_index in sheet_row_range(ws, first_candidate_row):
            for col_idx in metric_columns:
                value = read_candidate_cell(ws, row_index, col_idx)
                if not is_blank_value(value):
                    source_map[(sheet_name, row_index, col_idx)] = base_path.name
    return source_map


def copy_score_cells(
    *,
    base_workbook: Any,
    source_workbook: Any,
    source_path: Path,
    source_map: Dict[Tuple[str, int, int], str],
    score_sheets: Sequence[str],
    configured_metric_id_row: int,
    configured_first_candidate_row: int,
    metrics_by_id: Mapping[str, Mapping[str, Any]],
    roster_fields: Sequence[str],
) -> List[ScoreCellConflict]:
    conflicts = []
    for sheet_name in score_sheets:
        base_ws = base_workbook[sheet_name]
        source_ws = source_workbook[sheet_name]
        base_metric_id_row, base_metric_columns = metric_columns_for_matching_sheet(
            ws=base_ws,
            configured_metric_id_row=configured_metric_id_row,
            metrics_by_id=metrics_by_id,
            roster_fields=roster_fields,
        )
        source_metric_id_row, source_metric_columns = metric_columns_for_matching_sheet(
            ws=source_ws,
            configured_metric_id_row=configured_metric_id_row,
            metrics_by_id=metrics_by_id,
            roster_fields=roster_fields,
        )
        validate_matching_metric_columns(
            sheet_name=sheet_name,
            base_columns=base_metric_columns,
            source_columns=source_metric_columns,
            source_path=source_path,
        )
        first_candidate_row = max(
            configured_first_candidate_row,
            base_metric_id_row + 1,
            source_metric_id_row + 1,
        )

        for row_index in sheet_row_range(base_ws, first_candidate_row):
            for col_idx in base_metric_columns:
                incoming_value = read_candidate_cell(source_ws, row_index, col_idx)
                if is_blank_value(incoming_value):
                    continue

                base_cell = base_ws.cell(row=row_index, column=col_idx)
                existing_value = base_cell.value
                source_key = (sheet_name, row_index, col_idx)
                if is_blank_value(existing_value):
                    base_cell.value = incoming_value
                    source_map[source_key] = source_path.name
                    continue

                if not values_match(existing_value, incoming_value):
                    conflicts.append(
                        ScoreCellConflict(
                            sheet_name=sheet_name,
                            row_index=row_index,
                            column_index=col_idx,
                            existing_workbook=source_map.get(
                                source_key,
                                "base workbook",
                            ),
                            incoming_workbook=source_path.name,
                            existing_value=existing_value,
                            incoming_value=incoming_value,
                        )
                    )
    return conflicts


def format_conflict_message(conflicts: Sequence[ScoreCellConflict]) -> str:
    lines = ["Conflicting scores found while compressing node workbooks:"]
    for conflict in conflicts:
        lines.append(
            "- {sheet}!{cell} row {row}: {existing_workbook}={existing_value!r}, "
            "{incoming_workbook}={incoming_value!r}".format(
                sheet=conflict.sheet_name,
                cell=conflict.cell_ref,
                row=conflict.row_index,
                existing_workbook=conflict.existing_workbook,
                existing_value=conflict.existing_value,
                incoming_workbook=conflict.incoming_workbook,
                incoming_value=conflict.incoming_value,
            )
        )
    return "\n".join(lines)


def compress_template_workbooks(request: TemplateCompressionRequest) -> Path:
    workbook_paths = resolve_compression_workbooks(request)
    output_path = (
        request.output_path.resolve()
        if request.output_path
        else default_output_path(request)
    )
    if output_path in workbook_paths:
        raise ValueError("Output path must not overwrite an input workbook")

    loaded = load_generation_inputs(
        TemplateGenerationRequest(config_path=request.config_path)
    )
    sheet_contract = loaded.config_doc.get("sheet_contract", {})
    configured_metric_id_row = int(sheet_contract.get("metric_id_row", 2))
    configured_first_candidate_row = int(sheet_contract.get("first_candidate_row", 3))
    roster_fields = all_configured_roster_fields_from_template(loaded)

    base_path = workbook_paths[0]
    base_workbook = load_workbook(base_path)
    try:
        score_sheets = score_sheet_names(
            workbook=base_workbook,
            configured_metric_id_row=configured_metric_id_row,
            metrics_by_id=loaded.metrics_by_id,
            roster_fields=roster_fields,
        )
        if not score_sheets:
            raise ValueError("Base workbook has no score sheets: {0}".format(base_path))

        source_map = initialize_source_map(
            base_workbook=base_workbook,
            base_path=base_path,
            score_sheets=score_sheets,
            configured_metric_id_row=configured_metric_id_row,
            configured_first_candidate_row=configured_first_candidate_row,
            metrics_by_id=loaded.metrics_by_id,
            roster_fields=roster_fields,
        )

        conflicts = []
        for source_path in workbook_paths[1:]:
            source_workbook = load_workbook(source_path, data_only=True)
            try:
                validate_matching_score_sheets(
                    base_workbook=base_workbook,
                    source_workbook=source_workbook,
                    source_path=source_path,
                    configured_metric_id_row=configured_metric_id_row,
                    metrics_by_id=loaded.metrics_by_id,
                    roster_fields=roster_fields,
                )
                conflicts.extend(
                    copy_score_cells(
                        base_workbook=base_workbook,
                        source_workbook=source_workbook,
                        source_path=source_path,
                        source_map=source_map,
                        score_sheets=score_sheets,
                        configured_metric_id_row=configured_metric_id_row,
                        configured_first_candidate_row=configured_first_candidate_row,
                        metrics_by_id=loaded.metrics_by_id,
                        roster_fields=roster_fields,
                    )
                )
            finally:
                source_workbook.close()

        if conflicts:
            raise TemplateCompressionConflictError(conflicts)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        base_workbook.save(output_path)
    finally:
        base_workbook.close()

    return output_path


def main(argv: Optional[Sequence[str]] = None) -> int:
    try:
        output_path = compress_template_workbooks(
            request_from_namespace(parse_args(argv))
        )
    except TemplateCompressionConflictError as exc:
        print(str(exc))
        return 1
    print("Compressed node workbook generated: {0}".format(output_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
