from __future__ import annotations

import argparse
from copy import copy
from collections import OrderedDict, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, DefaultDict, Dict, Iterable, List, Mapping, MutableMapping, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from template_generator import (
    DEFAULT_CONFIG_PATH,
    HEADER_FILL,
    HEADER_FONT,
    LoadedTemplateConfig,
    TemplateGenerationRequest,
    add_min_pass_highlight_rule,
    apply_table_cell_style,
    configured_locked_left_columns,
    load_generation_inputs,
    load_yaml,
    resolve_path,
)


@dataclass(frozen=True)
class MasterGenerationRequest:
    config_path: Path
    workbook_paths: Sequence[Path] = ()
    output_path: Optional[Path] = None
    dropbox_path: Optional[Path] = None
    processed_path: Optional[Path] = None
    dynamic: bool = False


@dataclass(frozen=True)
class MasterWorkbookBatch:
    workbook_paths: Sequence[Path]
    pending_paths: Sequence[Path] = ()
    processed_path: Optional[Path] = None


@dataclass(frozen=True)
class LoadedMasterConfig:
    template: LoadedTemplateConfig
    master_path: Path
    master_doc: Dict[str, Any]


@dataclass
class CandidateAccumulator:
    source_values: DefaultDict[Tuple[str, str], List[Any]] = field(
        default_factory=lambda: defaultdict(list)
    )
    metric_values: DefaultDict[str, List[Any]] = field(
        default_factory=lambda: defaultdict(list)
    )


@dataclass
class EarlyExitAccumulator:
    source_values: DefaultDict[Tuple[str, str], List[Any]] = field(
        default_factory=lambda: defaultdict(list)
    )
    reasons: List[str] = field(default_factory=list)
    source_workbooks: List[str] = field(default_factory=list)
    sheet_names: List[str] = field(default_factory=list)


@dataclass(frozen=True)
class MasterBuildResult:
    rows: Sequence[OrderedDict[str, Any]]
    early_exit_rows: Sequence[OrderedDict[str, Any]]


@dataclass
class DynamicMetricStyle:
    header_fill: Any = None
    body_fill: Any = None
    number_format: str = "General"


@dataclass
class DynamicMasterBuildResult:
    rows: Sequence[OrderedDict[str, Any]]
    early_exit_rows: Sequence[OrderedDict[str, Any]]
    metric_ids: Sequence[str]
    metric_styles: Mapping[str, DynamicMetricStyle]


EARLY_EXIT_REASONS = ("MED PULL", "DOR", "ADMIN PULL")
DEFAULT_NULL_RULE = {
    "trim_strings": True,
    "blank_inputs_are_null": True,
    "null_literals": ["N/A", "NA", "NULL", "null"],
    "on_all_inputs_null": None,
}


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Aggregate one or more node workbooks into the master workbook."
    )
    parser.add_argument(
        "--config",
        default=DEFAULT_CONFIG_PATH,
        help="Path to canonical config file.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output xlsx path (default: workbooks/MasterWorkbook_v{version}.xlsx).",
    )
    parser.add_argument(
        "--dropbox",
        default=None,
        help=(
            "Directory containing newly scored node workbooks. Existing files in "
            "dropbox/processed are included for cumulative compilation."
        ),
    )
    parser.add_argument(
        "--processed",
        default=None,
        help="Processed archive directory (default: <dropbox>/processed).",
    )
    parser.add_argument(
        "--dynamic",
        action="store_true",
        help=(
            "Build a flattened master directly from workbook metric_ids instead "
            "of config/master-config.yaml."
        ),
    )
    parser.add_argument(
        "workbooks",
        nargs="*",
        help="Optional explicit node workbook xlsx files to aggregate.",
    )
    return parser.parse_args(argv)


def request_from_namespace(args: argparse.Namespace) -> MasterGenerationRequest:
    return MasterGenerationRequest(
        config_path=Path(args.config).resolve(),
        workbook_paths=[Path(value).resolve() for value in args.workbooks],
        output_path=Path(args.output).resolve() if args.output else None,
        dropbox_path=Path(args.dropbox).resolve() if args.dropbox else None,
        processed_path=Path(args.processed).resolve() if args.processed else None,
        dynamic=bool(args.dynamic),
    )


def validation_contract(config_doc: Mapping[str, Any]) -> Mapping[str, Any]:
    validation = config_doc.get("validation", {})
    if not isinstance(validation, dict):
        raise ValueError("config.yaml 'validation' must be a mapping")
    return validation


def validate_supported_value(
    *,
    label: str,
    value: Any,
    supported_values: Iterable[Any],
) -> None:
    if value not in set(supported_values):
        raise ValueError(
            "{0} uses unsupported value '{1}'".format(label, value)
        )


def validate_duplicate_rule(
    *,
    label: str,
    rule: Mapping[str, Any],
    validation: Mapping[str, Any],
) -> None:
    mode = rule.get("mode", "error_on_conflict")
    validate_supported_value(
        label="{0}.mode".format(label),
        value=mode,
        supported_values=validation.get("supported_duplicate_resolution_modes", []),
    )
    if mode == "aggregate":
        aggregate_function = rule.get("aggregate_function")
        if not aggregate_function:
            raise ValueError(
                "{0} requires aggregate_function when mode is 'aggregate'".format(label)
            )
        validate_supported_value(
            label="{0}.aggregate_function".format(label),
            value=aggregate_function,
            supported_values=validation.get("supported_aggregate_functions", []),
        )


def validate_null_rule(
    *,
    label: str,
    rule: Mapping[str, Any],
    validation: Mapping[str, Any],
) -> None:
    output_value = rule.get("on_all_inputs_null", "null")
    validate_supported_value(
        label="{0}.on_all_inputs_null".format(label),
        value=output_value,
        supported_values=validation.get("supported_null_outputs", []),
    )


def validate_master_config(loaded: LoadedMasterConfig) -> None:
    validation = validation_contract(loaded.template.config_doc)
    master_doc = loaded.master_doc

    columns = master_doc.get("columns")
    if not isinstance(columns, list) or not columns:
        raise ValueError("master-config.yaml must contain a non-empty 'columns' list")

    defaults = master_doc.get("defaults", {})
    if not isinstance(defaults, dict):
        raise ValueError("master-config.yaml 'defaults' must be a mapping")

    default_duplicate_rule = defaults.get("duplicate_metric_rule", {})
    if not isinstance(default_duplicate_rule, dict):
        raise ValueError(
            "master-config.yaml 'defaults.duplicate_metric_rule' must be a mapping"
        )
    validate_duplicate_rule(
        label="defaults.duplicate_metric_rule",
        rule=default_duplicate_rule,
        validation=validation,
    )

    default_null_rule = defaults.get("null_rule", {})
    if not isinstance(default_null_rule, dict):
        raise ValueError("master-config.yaml 'defaults.null_rule' must be a mapping")
    validate_null_rule(
        label="defaults.null_rule",
        rule=default_null_rule,
        validation=validation,
    )

    seen_column_ids = set()
    for column in columns:
        if not isinstance(column, dict):
            raise ValueError("Each master-config column must be a mapping")

        column_id = column.get("column_id")
        if not column_id:
            raise ValueError("Master column missing column_id")
        if column_id in seen_column_ids:
            raise ValueError("Duplicate master column_id: {0}".format(column_id))
        seen_column_ids.add(column_id)

        source = column.get("source")
        if not isinstance(source, dict):
            raise ValueError(
                "Master column '{0}' is missing a valid source mapping".format(column_id)
            )

        transform = column.get("transform", {})
        if not isinstance(transform, dict):
            raise ValueError(
                "Master column '{0}' transform must be a mapping".format(column_id)
            )

        transform_op = transform.get("op", "identity")
        validate_supported_value(
            label="columns.{0}.transform.op".format(column_id),
            value=transform_op,
            supported_values=validation.get("supported_transform_ops", []),
        )

        source_kind = source.get("kind")
        if source_kind == "metric":
            metric_ids = source.get("metric_ids")
            if not isinstance(metric_ids, list) or not metric_ids:
                raise ValueError(
                    "Master column '{0}' metric source must define metric_ids".format(
                        column_id
                    )
                )
            for metric_id in metric_ids:
                if metric_id not in loaded.template.metrics_by_id:
                    raise ValueError(
                        "Master column '{0}' references unknown metric_id '{1}'".format(
                            column_id, metric_id
                        )
                    )
        elif source_kind in {"roster_field", "meta_field", "ingestion_metadata"}:
            if not source.get("field"):
                raise ValueError(
                    "Master column '{0}' source kind '{1}' requires a field".format(
                        column_id, source_kind
                    )
                )
        else:
            raise ValueError(
                "Master column '{0}' uses unsupported source kind '{1}'".format(
                    column_id, source_kind
                )
            )

        duplicate_rule = transform.get("duplicate_metric_rule", default_duplicate_rule)
        if not isinstance(duplicate_rule, dict):
            raise ValueError(
                "Master column '{0}' duplicate_metric_rule must be a mapping".format(
                    column_id
                )
            )
        validate_duplicate_rule(
            label="columns.{0}.transform.duplicate_metric_rule".format(column_id),
            rule=duplicate_rule,
            validation=validation,
        )

        null_rule = column.get("null_rule", default_null_rule)
        if not isinstance(null_rule, dict):
            raise ValueError(
                "Master column '{0}' null_rule must be a mapping".format(column_id)
            )
        validate_null_rule(
            label="columns.{0}.null_rule".format(column_id),
            rule=null_rule,
            validation=validation,
        )


def load_master_generation_inputs(
    request: MasterGenerationRequest,
) -> LoadedMasterConfig:
    template = load_generation_inputs(
        TemplateGenerationRequest(
            config_path=request.config_path,
            entry_rows=1,
        )
    )

    files = template.config_doc.get("files", {})
    if not isinstance(files, dict):
        raise ValueError("config.yaml 'files' must be a mapping")

    master_candidate = files.get("master")
    if not master_candidate:
        raise ValueError("config.yaml 'files.master' is required for master generation")
    master_path = resolve_path(template.config_path, master_candidate)
    master_doc = load_yaml(master_path)

    loaded = LoadedMasterConfig(
        template=template,
        master_path=master_path,
        master_doc=master_doc,
    )
    validate_master_config(loaded)
    return loaded


def default_output_path(loaded: LoadedMasterConfig) -> Path:
    return (
        loaded.template.config_path.parent.parent
        / "workbooks"
        / "MasterWorkbook_v{0}.xlsx".format(loaded.master_doc.get("version", "unknown"))
    ).resolve()


def default_dynamic_output_path(template: LoadedTemplateConfig) -> Path:
    return (
        template.config_path.parent.parent
        / "workbooks"
        / "DynamicMasterWorkbook_v{0}.xlsx".format(
            template.config_doc.get("version", "unknown")
        )
    ).resolve()


def is_excel_workbook(path: Path) -> bool:
    return (
        path.is_file()
        and path.suffix.lower() == ".xlsx"
        and not path.name.startswith("~$")
    )


def discover_workbooks(directory: Path) -> List[Path]:
    if not directory.exists():
        return []
    if not directory.is_dir():
        raise ValueError("Workbook source is not a directory: {0}".format(directory))
    return sorted(
        (path.resolve() for path in directory.iterdir() if is_excel_workbook(path)),
        key=lambda path: path.name,
    )


def dedupe_paths(paths: Sequence[Path]) -> List[Path]:
    deduped = []  # type: List[Path]
    seen = set()
    for path in paths:
        resolved = path.resolve()
        if resolved not in seen:
            seen.add(resolved)
            deduped.append(resolved)
    return deduped


def validate_explicit_workbook_paths(workbook_paths: Sequence[Path]) -> None:
    for workbook_path in workbook_paths:
        if not is_excel_workbook(workbook_path):
            raise ValueError(
                "Node workbook path is not a readable .xlsx file: {0}".format(
                    workbook_path
                )
            )


def path_contains(parent: Path, child: Path) -> bool:
    parent = parent.resolve()
    child = child.resolve()
    return child == parent or child.is_relative_to(parent)


def validate_dropbox_output_path(
    *,
    output_path: Path,
    dropbox_path: Path,
    processed_path: Path,
) -> None:
    if path_contains(dropbox_path, output_path) or path_contains(
        processed_path, output_path
    ):
        raise ValueError(
            "Master output must not be written inside the dropbox or processed folder"
        )


def resolve_master_workbook_batch(
    request: MasterGenerationRequest,
    *,
    output_path: Path,
) -> MasterWorkbookBatch:
    explicit_paths = dedupe_paths(
        [Path(path).resolve() for path in request.workbook_paths]
    )
    validate_explicit_workbook_paths(explicit_paths)

    if request.processed_path and request.dropbox_path is None:
        raise ValueError("--processed requires --dropbox")

    if request.dropbox_path is None:
        if not explicit_paths:
            raise ValueError("Provide at least one workbook path or --dropbox")
        return MasterWorkbookBatch(workbook_paths=explicit_paths)

    dropbox_path = request.dropbox_path.resolve()
    if not dropbox_path.is_dir():
        raise ValueError("Dropbox path is not a directory: {0}".format(dropbox_path))

    processed_path = (
        request.processed_path.resolve()
        if request.processed_path
        else (dropbox_path / "processed").resolve()
    )
    if processed_path.exists() and not processed_path.is_dir():
        raise ValueError(
            "Processed path is not a directory: {0}".format(processed_path)
        )

    validate_dropbox_output_path(
        output_path=output_path,
        dropbox_path=dropbox_path,
        processed_path=processed_path,
    )

    pending_paths = discover_workbooks(dropbox_path)
    processed_paths = discover_workbooks(processed_path)

    for pending_path in pending_paths:
        target_path = processed_path / pending_path.name
        if target_path.exists():
            raise ValueError(
                "Processed workbook already exists for dropbox file '{0}': {1}".format(
                    pending_path.name,
                    target_path,
                )
            )

    workbook_paths = dedupe_paths(
        list(explicit_paths) + processed_paths + pending_paths
    )
    if not workbook_paths:
        raise ValueError("No .xlsx node workbooks found to compile")

    return MasterWorkbookBatch(
        workbook_paths=workbook_paths,
        pending_paths=pending_paths,
        processed_path=processed_path,
    )


def move_processed_workbooks(batch: MasterWorkbookBatch) -> None:
    if not batch.pending_paths:
        return
    if batch.processed_path is None:
        raise ValueError("Internal master generation state is missing processed path")

    batch.processed_path.mkdir(parents=True, exist_ok=True)
    for pending_path in batch.pending_paths:
        pending_path.replace(batch.processed_path / pending_path.name)


def normalize_null_like(value: Any, rule: Mapping[str, Any]) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        candidate = value.strip() if rule.get("trim_strings", True) else value
        if candidate == "" and rule.get("blank_inputs_are_null", True):
            return None
        if candidate in set(rule.get("null_literals", [])):
            return None
        return candidate
    return value


def metric_entry_style(metric: Mapping[str, Any]) -> Optional[str]:
    excel_input = metric.get("excel_input", {})
    if not isinstance(excel_input, dict):
        return None
    entry_style = excel_input.get("entry_style")
    return str(entry_style) if entry_style else None


def normalize_mm_ss_decimal(value: Decimal) -> Optional[float]:
    if value < 0:
        return None

    minutes = int(value)
    seconds_decimal = (value - Decimal(minutes)) * Decimal(100)
    if seconds_decimal != seconds_decimal.to_integral_value():
        return None

    seconds = int(seconds_decimal)
    if seconds >= 60:
        return None
    return float(minutes * 60 + seconds)


def normalize_mm_ss_value(value: Any) -> Optional[float]:
    try:
        decimal_value = Decimal(str(value).strip())
    except (InvalidOperation, AttributeError):
        return None
    return normalize_mm_ss_decimal(decimal_value)


def normalize_time_string(raw: str, entry_style: Optional[str] = None) -> Optional[float]:
    stripped = raw.strip()
    if not stripped:
        return None

    if entry_style == "mm_ss" and "." in stripped:
        parsed = normalize_mm_ss_value(stripped)
        if parsed is not None:
            return parsed

    parts = stripped.split(":")
    try:
        if len(parts) == 2:
            minutes = int(parts[0])
            seconds = float(parts[1])
            if seconds < 0 or seconds >= 60:
                return None
            return minutes * 60 + seconds
        if len(parts) == 3:
            hours = int(parts[0])
            minutes = int(parts[1])
            seconds = float(parts[2])
            if minutes < 0 or minutes >= 60 or seconds < 0 or seconds >= 60:
                return None
            return hours * 3600 + minutes * 60 + seconds
    except ValueError:
        return None
    return None


def normalize_timed_value(
    value: Any, entry_style: Optional[str] = None
) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, timedelta):
        total_seconds = value.total_seconds()
        if entry_style == "mm_ss" and abs(total_seconds) >= 86400:
            parsed = normalize_mm_ss_value(total_seconds / 86400)
            if parsed is not None:
                return parsed
        return total_seconds
    if isinstance(value, datetime):
        return (
            value.hour * 3600
            + value.minute * 60
            + value.second
            + value.microsecond / 1_000_000.0
        )
    if isinstance(value, time):
        return (
            value.hour * 3600
            + value.minute * 60
            + value.second
            + value.microsecond / 1_000_000.0
        )
    if isinstance(value, (int, float)):
        if entry_style == "mm_ss":
            parsed = normalize_mm_ss_value(value)
            if parsed is not None:
                return parsed
        return float(value) * 86400 if abs(float(value)) <= 2 else float(value)
    if isinstance(value, str):
        parsed = normalize_time_string(value, entry_style=entry_style)
        if parsed is not None:
            return parsed
    raise ValueError("Unsupported timed value '{0}'".format(value))


def normalize_numeric_value(value: Any, *, integer: bool) -> Optional[Any]:
    if value is None:
        return None
    if isinstance(value, bool):
        numeric = int(value)
    elif isinstance(value, (int, float)):
        numeric = float(value)
    elif isinstance(value, str):
        numeric = float(value.strip())
    else:
        raise ValueError("Unsupported numeric value '{0}'".format(value))

    if integer:
        return int(round(numeric))
    return numeric


def normalize_metric_value(
    *,
    raw_value: Any,
    metric: Mapping[str, Any],
    null_rule: Mapping[str, Any],
) -> Any:
    value = normalize_null_like(raw_value, null_rule)
    if value is None:
        return None

    metric_type = metric.get("type")
    if metric_type == "timed":
        return normalize_timed_value(value, entry_style=metric_entry_style(metric))
    if metric_type == "integer":
        return normalize_numeric_value(value, integer=True)
    if metric_type == "numeric":
        return normalize_numeric_value(value, integer=False)
    if metric_type in {"categorical", "text"}:
        return str(value)
    return value


def is_numeric_metric(metric: Mapping[str, Any]) -> bool:
    return metric.get("type") in {"integer", "numeric", "timed"}


def is_subjective_metric(metric: Mapping[str, Any]) -> bool:
    return metric.get("domain_ref") == "subjective"


def metric_display_name(metric_id: str, metric: Mapping[str, Any]) -> str:
    return str(metric.get("display_name", metric_id))


def append_unique(values: List[str], value: str) -> None:
    if value not in values:
        values.append(value)


def capture_dynamic_metric_style(
    *,
    metric_styles: MutableMapping[str, DynamicMetricStyle],
    metric_id: str,
    ws: Any,
    metric_id_row: int,
    first_candidate_row: int,
    col_idx: int,
) -> None:
    if metric_id in metric_styles:
        return

    header_row = max(1, metric_id_row - 2)
    metric_styles[metric_id] = DynamicMetricStyle(
        header_fill=copy(ws.cell(row=header_row, column=col_idx).fill),
        body_fill=copy(ws.cell(row=first_candidate_row, column=col_idx).fill),
        number_format=ws.cell(row=first_candidate_row, column=col_idx).number_format,
    )


def effective_dynamic_raw_values(
    *,
    raw_values: Sequence[Any],
    metric: Mapping[str, Any],
    null_rule: Mapping[str, Any],
) -> Sequence[Any]:
    if not is_subjective_metric(metric):
        return raw_values

    has_score = any(
        normalize_null_like(value, null_rule) is not None for value in raw_values
    )
    if not has_score:
        return []
    return [
        2 if normalize_null_like(value, null_rule) is None else value
        for value in raw_values
    ]


def resolve_dynamic_metric_value(
    *,
    values: Sequence[Any],
    metric: Mapping[str, Any],
) -> Any:
    non_null = [value for value in values if value is not None]
    if not non_null:
        return None

    if is_numeric_metric(metric):
        numeric_values = [float(value) for value in non_null]
        average = sum(numeric_values) / len(numeric_values)
        return int(average) if average.is_integer() else average

    distinct_values = dedupe_preserve_order(non_null)
    return "; ".join(str(value) for value in distinct_values)


def read_meta_sheet(wb: Any) -> Dict[str, Any]:
    if "META" not in wb.sheetnames:
        raise ValueError("Workbook is missing required META sheet")

    ws = wb["META"]
    meta = {}
    row_index = 1
    while True:
        key = ws.cell(row=row_index, column=1).value
        if key in (None, ""):
            break
        meta[str(key)] = ws.cell(row=row_index, column=2).value
        row_index += 1
    return meta


def validate_node_workbook_meta(
    *,
    workbook_path: Path,
    meta: Mapping[str, Any],
    loaded: LoadedMasterConfig,
) -> None:
    expected_registry = loaded.template.config_doc.get("registry_name")
    if meta.get("registry_name") != expected_registry:
        raise ValueError(
            "Workbook {0} uses registry_name '{1}', expected '{2}'".format(
                workbook_path,
                meta.get("registry_name"),
                expected_registry,
            )
        )

    expected_version = str(loaded.template.config_doc.get("version"))
    actual_version = str(meta.get("config_version"))
    if actual_version != expected_version:
        raise ValueError(
            "Workbook {0} uses config_version '{1}', expected '{2}'".format(
                workbook_path,
                actual_version,
                expected_version,
            )
        )


def configured_roster_fields(loaded: LoadedMasterConfig) -> List[str]:
    sheet_contract = loaded.template.config_doc.get("sheet_contract", {})
    return configured_locked_left_columns(sheet_contract)


def all_configured_roster_fields(loaded: LoadedMasterConfig) -> List[str]:
    sheet_contract = loaded.template.config_doc.get("sheet_contract", {})
    roster_fields = configured_locked_left_columns(sheet_contract)
    for evolution in loaded.template.evolutions_doc.get("evolutions", []):
        if not isinstance(evolution, dict):
            continue
        for field_name in configured_locked_left_columns(sheet_contract, evolution):
            if field_name not in roster_fields:
                roster_fields.append(field_name)
    return roster_fields


def fallback_roster_row(
    *,
    loaded: LoadedMasterConfig,
    row_index: int,
    first_candidate_row: int,
) -> Mapping[str, Any]:
    roster_index = row_index - first_candidate_row
    if 0 <= roster_index < len(loaded.template.roster_rows):
        return loaded.template.roster_rows[roster_index]
    return {}


def read_candidate_roster_value(
    *,
    ws: Any,
    row_index: int,
    field_name: str,
    roster_columns: Mapping[str, int],
    fallback_row: Mapping[str, Any],
) -> Any:
    column_index = roster_columns.get(field_name)
    if column_index is not None:
        return read_candidate_cell(ws, row_index, column_index)
    return fallback_row.get(field_name)


def sheet_row_range(ws: Any, start_row: int) -> range:
    return range(start_row, ws.max_row + 1)


def read_candidate_cell(ws: Any, row_index: int, column_index: int) -> Any:
    return ws.cell(row=row_index, column=column_index).value


def source_key(kind: str, field_name: str) -> Tuple[str, str]:
    return (kind, field_name)


def accumulate_source_value(
    *,
    bucket: CandidateAccumulator | EarlyExitAccumulator,
    kind: str,
    field_name: str,
    value: Any,
    null_rule: Mapping[str, Any],
) -> None:
    bucket.source_values[source_key(kind, field_name)].append(
        normalize_null_like(value, null_rule)
    )


def available_source_roster_fields(roster_fields: Sequence[str]) -> List[str]:
    available_roster_fields = []
    for field_name in ["uid", "first", "last"] + list(roster_fields):
        if field_name not in available_roster_fields:
            available_roster_fields.append(field_name)
    return available_roster_fields


def accumulate_candidate_source_values(
    *,
    bucket: CandidateAccumulator | EarlyExitAccumulator,
    ws: Any,
    row_index: int,
    roster_columns: Mapping[str, int],
    roster_fallback: Mapping[str, Any],
    roster_fields: Sequence[str],
    meta: Mapping[str, Any],
    workbook_path: Path,
    null_rule: Mapping[str, Any],
) -> None:
    for field_name in available_source_roster_fields(roster_fields):
        accumulate_source_value(
            bucket=bucket,
            kind="roster_field",
            field_name=field_name,
            value=read_candidate_roster_value(
                ws=ws,
                row_index=row_index,
                field_name=field_name,
                roster_columns=roster_columns,
                fallback_row=roster_fallback,
            ),
            null_rule=null_rule,
        )
    accumulate_source_value(
        bucket=bucket,
        kind="meta_field",
        field_name="block_number",
        value=meta.get("block_number"),
        null_rule=null_rule,
    )
    accumulate_source_value(
        bucket=bucket,
        kind="meta_field",
        field_name="fiscal_year",
        value=meta.get("fiscal_year"),
        null_rule=null_rule,
    )
    accumulate_source_value(
        bucket=bucket,
        kind="ingestion_metadata",
        field_name="source_workbook",
        value=workbook_path.name,
        null_rule=null_rule,
    )


def metric_columns_for_sheet(
    *,
    ws: Any,
    metric_id_row: int,
    metrics_by_id: Mapping[str, Mapping[str, Any]],
) -> Dict[int, str]:
    metric_columns = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=metric_id_row, column=col_idx).value
        if value in metrics_by_id:
            metric_columns[col_idx] = str(value)
    return metric_columns


def normalize_early_exit_reason(value: Any) -> Optional[str]:
    if not isinstance(value, str):
        return None
    normalized = " ".join(value.strip().upper().split())
    if normalized in EARLY_EXIT_REASONS:
        return normalized
    return None


def early_exit_reasons_for_row(
    *,
    ws: Any,
    row_index: int,
    metric_columns: Mapping[int, str],
) -> List[str]:
    reasons = []
    for col_idx in metric_columns:
        reason = normalize_early_exit_reason(
            read_candidate_cell(ws, row_index, col_idx)
        )
        if reason is not None and reason not in reasons:
            reasons.append(reason)
    return reasons


def record_early_exit(
    *,
    early_exits: MutableMapping[Tuple[str, str, str], EarlyExitAccumulator],
    key: Tuple[str, str, str],
    reasons: Sequence[str],
    ws: Any,
    row_index: int,
    roster_columns: Mapping[str, int],
    roster_fallback: Mapping[str, Any],
    roster_fields: Sequence[str],
    meta: Mapping[str, Any],
    workbook_path: Path,
    sheet_name: str,
    null_rule: Mapping[str, Any],
) -> None:
    bucket = early_exits.setdefault(key, EarlyExitAccumulator())
    accumulate_candidate_source_values(
        bucket=bucket,
        ws=ws,
        row_index=row_index,
        roster_columns=roster_columns,
        roster_fallback=roster_fallback,
        roster_fields=roster_fields,
        meta=meta,
        workbook_path=workbook_path,
        null_rule=null_rule,
    )
    for reason in reasons:
        if reason not in bucket.reasons:
            bucket.reasons.append(reason)
    if workbook_path.name not in bucket.source_workbooks:
        bucket.source_workbooks.append(workbook_path.name)
    if sheet_name not in bucket.sheet_names:
        bucket.sheet_names.append(sheet_name)


def roster_columns_for_sheet(
    *,
    ws: Any,
    metric_id_row: int,
    roster_fields: Sequence[str],
) -> Dict[str, int]:
    roster_columns = {}
    roster_field_set = set(roster_fields)
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=metric_id_row, column=col_idx).value
        if value in roster_field_set:
            roster_columns[str(value)] = col_idx
    return roster_columns


def detect_machine_header_row(
    *,
    ws: Any,
    configured_metric_id_row: int,
    metrics_by_id: Mapping[str, Mapping[str, Any]],
    roster_fields: Sequence[str],
) -> int:
    rows_to_check = [configured_metric_id_row]
    for row_index in range(1, min(ws.max_row, configured_metric_id_row + 4) + 1):
        if row_index not in rows_to_check:
            rows_to_check.append(row_index)
    roster_field_set = set(roster_fields)

    for row_index in rows_to_check:
        has_metric_header = False
        has_roster_header = False
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=row_index, column=col_idx).value
            if value in metrics_by_id:
                has_metric_header = True
            if value in roster_field_set:
                has_roster_header = True
        if has_metric_header and has_roster_header:
            return row_index

    return configured_metric_id_row


def detected_first_candidate_row(
    *,
    configured_first_candidate_row: int,
    machine_header_row: int,
) -> int:
    return max(configured_first_candidate_row, machine_header_row + 1)


def ingest_node_workbook(
    *,
    workbook_path: Path,
    loaded: LoadedMasterConfig,
    buckets: MutableMapping[Tuple[str, str, str], CandidateAccumulator],
    early_exits: MutableMapping[Tuple[str, str, str], EarlyExitAccumulator],
) -> None:
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        meta = read_meta_sheet(workbook)
        validate_node_workbook_meta(
            workbook_path=workbook_path,
            meta=meta,
            loaded=loaded,
        )

        default_null_rule = loaded.master_doc.get("defaults", {}).get("null_rule", {})
        roster_fields = all_configured_roster_fields(loaded)
        sheet_contract = loaded.template.config_doc.get("sheet_contract", {})
        configured_metric_id_row = int(sheet_contract.get("metric_id_row", 2))
        configured_first_candidate_row = int(
            sheet_contract.get("first_candidate_row", 3)
        )

        system_sheets = {"META", "ROSTER", "LOOKUPS", "MASTER"}
        for sheet_name in workbook.sheetnames:
            if sheet_name in system_sheets:
                continue

            ws = workbook[sheet_name]
            metric_id_row = detect_machine_header_row(
                ws=ws,
                configured_metric_id_row=configured_metric_id_row,
                metrics_by_id=loaded.template.metrics_by_id,
                roster_fields=roster_fields,
            )
            first_candidate_row = detected_first_candidate_row(
                configured_first_candidate_row=configured_first_candidate_row,
                machine_header_row=metric_id_row,
            )
            metric_columns = metric_columns_for_sheet(
                ws=ws,
                metric_id_row=metric_id_row,
                metrics_by_id=loaded.template.metrics_by_id,
            )
            if not metric_columns:
                continue

            roster_columns = roster_columns_for_sheet(
                ws=ws,
                metric_id_row=metric_id_row,
                roster_fields=roster_fields,
            )
            for row_index in sheet_row_range(ws, first_candidate_row):
                roster_fallback = fallback_roster_row(
                    loaded=loaded,
                    row_index=row_index,
                    first_candidate_row=first_candidate_row,
                )
                uid_value = normalize_null_like(
                    read_candidate_roster_value(
                        ws=ws,
                        row_index=row_index,
                        field_name="uid",
                        roster_columns=roster_columns,
                        fallback_row=roster_fallback,
                    ),
                    default_null_rule,
                )
                if uid_value is None:
                    continue

                key = (
                    str(uid_value),
                    str(meta.get("block_number", "")),
                    str(meta.get("fiscal_year", "")),
                )

                early_exit_reasons = early_exit_reasons_for_row(
                    ws=ws,
                    row_index=row_index,
                    metric_columns=metric_columns,
                )
                if early_exit_reasons:
                    buckets.pop(key, None)
                    record_early_exit(
                        early_exits=early_exits,
                        key=key,
                        reasons=early_exit_reasons,
                        ws=ws,
                        row_index=row_index,
                        roster_columns=roster_columns,
                        roster_fallback=roster_fallback,
                        roster_fields=roster_fields,
                        meta=meta,
                        workbook_path=workbook_path,
                        sheet_name=sheet_name,
                        null_rule=default_null_rule,
                    )
                    continue
                if key in early_exits:
                    continue

                bucket = buckets.setdefault(key, CandidateAccumulator())
                accumulate_candidate_source_values(
                    bucket=bucket,
                    ws=ws,
                    row_index=row_index,
                    roster_columns=roster_columns,
                    roster_fallback=roster_fallback,
                    roster_fields=roster_fields,
                    meta=meta,
                    workbook_path=workbook_path,
                    null_rule=default_null_rule,
                )

                for col_idx, metric_id in metric_columns.items():
                    raw_value = read_candidate_cell(ws, row_index, col_idx)
                    normalized_value = normalize_metric_value(
                        raw_value=raw_value,
                        metric=loaded.template.metrics_by_id[metric_id],
                        null_rule=default_null_rule,
                    )
                    bucket.metric_values[metric_id].append(normalized_value)
    finally:
        workbook.close()


def ingest_node_workbook_dynamic(
    *,
    workbook_path: Path,
    loaded: LoadedMasterConfig,
    buckets: MutableMapping[Tuple[str, str, str], CandidateAccumulator],
    early_exits: MutableMapping[Tuple[str, str, str], EarlyExitAccumulator],
    metric_ids: List[str],
    metric_styles: MutableMapping[str, DynamicMetricStyle],
) -> None:
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        meta = read_meta_sheet(workbook)
        validate_node_workbook_meta(
            workbook_path=workbook_path,
            meta=meta,
            loaded=loaded,
        )

        default_null_rule = DEFAULT_NULL_RULE
        roster_fields = all_configured_roster_fields(loaded)
        sheet_contract = loaded.template.config_doc.get("sheet_contract", {})
        configured_metric_id_row = int(sheet_contract.get("metric_id_row", 2))
        configured_first_candidate_row = int(
            sheet_contract.get("first_candidate_row", 3)
        )

        system_sheets = {"META", "ROSTER", "LOOKUPS", "MASTER", "EARLY_EXITS"}
        for sheet_name in workbook.sheetnames:
            if sheet_name in system_sheets:
                continue

            ws = workbook[sheet_name]
            metric_id_row = detect_machine_header_row(
                ws=ws,
                configured_metric_id_row=configured_metric_id_row,
                metrics_by_id=loaded.template.metrics_by_id,
                roster_fields=roster_fields,
            )
            first_candidate_row = detected_first_candidate_row(
                configured_first_candidate_row=configured_first_candidate_row,
                machine_header_row=metric_id_row,
            )
            metric_columns = metric_columns_for_sheet(
                ws=ws,
                metric_id_row=metric_id_row,
                metrics_by_id=loaded.template.metrics_by_id,
            )
            if not metric_columns:
                continue

            for col_idx, metric_id in metric_columns.items():
                append_unique(metric_ids, metric_id)
                capture_dynamic_metric_style(
                    metric_styles=metric_styles,
                    metric_id=metric_id,
                    ws=ws,
                    metric_id_row=metric_id_row,
                    first_candidate_row=first_candidate_row,
                    col_idx=col_idx,
                )

            roster_columns = roster_columns_for_sheet(
                ws=ws,
                metric_id_row=metric_id_row,
                roster_fields=roster_fields,
            )
            for row_index in sheet_row_range(ws, first_candidate_row):
                roster_fallback = fallback_roster_row(
                    loaded=loaded,
                    row_index=row_index,
                    first_candidate_row=first_candidate_row,
                )
                uid_value = normalize_null_like(
                    read_candidate_roster_value(
                        ws=ws,
                        row_index=row_index,
                        field_name="uid",
                        roster_columns=roster_columns,
                        fallback_row=roster_fallback,
                    ),
                    default_null_rule,
                )
                if uid_value is None:
                    continue

                key = (
                    str(uid_value),
                    str(meta.get("block_number", "")),
                    str(meta.get("fiscal_year", "")),
                )

                early_exit_reasons = early_exit_reasons_for_row(
                    ws=ws,
                    row_index=row_index,
                    metric_columns=metric_columns,
                )
                if early_exit_reasons:
                    buckets.pop(key, None)
                    record_early_exit(
                        early_exits=early_exits,
                        key=key,
                        reasons=early_exit_reasons,
                        ws=ws,
                        row_index=row_index,
                        roster_columns=roster_columns,
                        roster_fallback=roster_fallback,
                        roster_fields=roster_fields,
                        meta=meta,
                        workbook_path=workbook_path,
                        sheet_name=sheet_name,
                        null_rule=default_null_rule,
                    )
                    continue
                if key in early_exits:
                    continue

                bucket = buckets.setdefault(key, CandidateAccumulator())
                accumulate_candidate_source_values(
                    bucket=bucket,
                    ws=ws,
                    row_index=row_index,
                    roster_columns=roster_columns,
                    roster_fallback=roster_fallback,
                    roster_fields=roster_fields,
                    meta=meta,
                    workbook_path=workbook_path,
                    null_rule=default_null_rule,
                )

                row_metric_values = defaultdict(list)
                for col_idx, metric_id in metric_columns.items():
                    row_metric_values[metric_id].append(
                        read_candidate_cell(ws, row_index, col_idx)
                    )

                for metric_id, raw_values in row_metric_values.items():
                    metric = loaded.template.metrics_by_id[metric_id]
                    for raw_value in effective_dynamic_raw_values(
                        raw_values=raw_values,
                        metric=metric,
                        null_rule=default_null_rule,
                    ):
                        normalized_value = normalize_metric_value(
                            raw_value=raw_value,
                            metric=metric,
                            null_rule=default_null_rule,
                        )
                        if normalized_value is not None:
                            bucket.metric_values[metric_id].append(normalized_value)
    finally:
        workbook.close()


def render_null_output(rule: Mapping[str, Any]) -> Any:
    mode = rule.get("on_all_inputs_null", "null")
    if mode in (None, "null"):
        return None
    if mode == "blank":
        return ""
    if mode == "zero":
        return 0
    if mode == "literal":
        return rule.get("literal", "")
    raise ValueError("Unsupported null output mode '{0}'".format(mode))


def dedupe_preserve_order(values: Sequence[Any]) -> List[Any]:
    seen = []
    for value in values:
        if value not in seen:
            seen.append(value)
    return seen


def apply_aggregate_function(values: Sequence[Any], aggregate_function: str) -> Any:
    if aggregate_function == "count":
        return len(values)

    numeric_values = [float(value) for value in values]
    if aggregate_function == "average":
        return sum(numeric_values) / len(numeric_values)
    if aggregate_function == "sum":
        return sum(numeric_values)
    if aggregate_function == "min":
        return min(numeric_values)
    if aggregate_function == "max":
        return max(numeric_values)
    raise ValueError(
        "Unsupported aggregate function '{0}'".format(aggregate_function)
    )


def resolve_values(
    *,
    values: Sequence[Any],
    duplicate_rule: Mapping[str, Any],
    null_rule: Mapping[str, Any],
    column_id: str,
) -> Any:
    non_null = [value for value in values if value is not None]
    if not non_null:
        return render_null_output(null_rule)

    distinct_values = dedupe_preserve_order(non_null)
    if len(distinct_values) == 1:
        return distinct_values[0]

    mode = duplicate_rule.get("mode", "error_on_conflict")
    if mode == "error_on_conflict":
        raise ValueError(
            "Column '{0}' encountered conflicting values: {1}".format(
                column_id, distinct_values
            )
        )
    if mode == "first_non_null":
        return non_null[0]
    if mode == "last_non_null":
        return non_null[-1]
    if mode == "concat_distinct":
        separator = duplicate_rule.get("separator", "; ")
        return separator.join(str(value) for value in distinct_values)
    if mode == "aggregate":
        return apply_aggregate_function(
            distinct_values, duplicate_rule.get("aggregate_function", "")
        )

    raise ValueError(
        "Column '{0}' uses unsupported duplicate mode '{1}'".format(column_id, mode)
    )


def evaluate_column(
    *,
    bucket: CandidateAccumulator,
    column: Mapping[str, Any],
    default_duplicate_rule: Mapping[str, Any],
    default_null_rule: Mapping[str, Any],
) -> Any:
    source = column.get("source", {})
    transform = column.get("transform", {})
    duplicate_rule = transform.get("duplicate_metric_rule", default_duplicate_rule)
    null_rule = column.get("null_rule", default_null_rule)
    column_id = str(column.get("column_id", "unknown"))

    source_kind = source.get("kind")
    if source_kind == "metric":
        values = []
        for metric_id in source.get("metric_ids", []):
            values.extend(bucket.metric_values.get(metric_id, []))
    else:
        field_name = str(source.get("field"))
        values = bucket.source_values.get(source_key(source_kind, field_name), [])

    return resolve_values(
        values=values,
        duplicate_rule=duplicate_rule,
        null_rule=null_rule,
        column_id=column_id,
    )


def resolve_early_exit_source_value(
    *,
    bucket: CandidateAccumulator | EarlyExitAccumulator,
    kind: str,
    field_name: str,
    default_null_rule: Mapping[str, Any],
) -> Any:
    return resolve_values(
        values=bucket.source_values.get(source_key(kind, field_name), []),
        duplicate_rule={"mode": "concat_distinct", "separator": "; "},
        null_rule=default_null_rule,
        column_id=field_name,
    )


def build_early_exit_rows(
    *,
    early_exits: Mapping[Tuple[str, str, str], EarlyExitAccumulator],
    default_null_rule: Mapping[str, Any],
) -> List[OrderedDict[str, Any]]:
    rows = []
    for bucket in early_exits.values():
        row = OrderedDict()
        row["ID"] = resolve_early_exit_source_value(
            bucket=bucket,
            kind="roster_field",
            field_name="uid",
            default_null_rule=default_null_rule,
        )
        row["Last Name"] = resolve_early_exit_source_value(
            bucket=bucket,
            kind="roster_field",
            field_name="last",
            default_null_rule=default_null_rule,
        )
        row["First Name"] = resolve_early_exit_source_value(
            bucket=bucket,
            kind="roster_field",
            field_name="first",
            default_null_rule=default_null_rule,
        )
        row["Block"] = resolve_early_exit_source_value(
            bucket=bucket,
            kind="meta_field",
            field_name="block_number",
            default_null_rule=default_null_rule,
        )
        row["Fiscal Year"] = resolve_early_exit_source_value(
            bucket=bucket,
            kind="meta_field",
            field_name="fiscal_year",
            default_null_rule=default_null_rule,
        )
        row["Exit Reason"] = "; ".join(bucket.reasons)
        row["Source"] = "; ".join(bucket.source_workbooks)
        row["Sheet"] = "; ".join(bucket.sheet_names)
        rows.append(row)
    return rows


def build_master_result(loaded: LoadedMasterConfig) -> MasterBuildResult:
    buckets: MutableMapping[Tuple[str, str, str], CandidateAccumulator] = OrderedDict()
    early_exits: MutableMapping[Tuple[str, str, str], EarlyExitAccumulator] = (
        OrderedDict()
    )
    for workbook_path in loaded_request_workbooks(loaded):
        ingest_node_workbook(
            workbook_path=workbook_path,
            loaded=loaded,
            buckets=buckets,
            early_exits=early_exits,
        )

    columns = loaded.master_doc.get("columns", [])
    defaults = loaded.master_doc.get("defaults", {})
    default_duplicate_rule = defaults.get("duplicate_metric_rule", {})
    default_null_rule = defaults.get("null_rule", {})

    rows = []
    for bucket in buckets.values():
        row = OrderedDict()
        for column in columns:
            row[str(column.get("header", column.get("column_id")))] = evaluate_column(
                bucket=bucket,
                column=column,
                default_duplicate_rule=default_duplicate_rule,
                default_null_rule=default_null_rule,
            )
        rows.append(row)
    early_exit_rows = build_early_exit_rows(
        early_exits=early_exits,
        default_null_rule=default_null_rule,
    )
    return MasterBuildResult(rows=rows, early_exit_rows=early_exit_rows)


def build_master_rows(loaded: LoadedMasterConfig) -> List[OrderedDict[str, Any]]:
    return list(build_master_result(loaded).rows)


def dynamic_metric_headers(
    *,
    metric_ids: Sequence[str],
    metrics_by_id: Mapping[str, Mapping[str, Any]],
) -> List[str]:
    display_names = [
        metric_display_name(metric_id, metrics_by_id[metric_id])
        for metric_id in metric_ids
    ]
    duplicate_names = {
        display_name
        for display_name in display_names
        if display_names.count(display_name) > 1
    }
    headers = []
    for metric_id, display_name in zip(metric_ids, display_names):
        if display_name in duplicate_names:
            headers.append("{0} ({1})".format(display_name, metric_id))
        else:
            headers.append(display_name)
    return headers


def build_dynamic_master_result(
    *,
    template: LoadedTemplateConfig,
    workbook_paths: Sequence[Path],
) -> DynamicMasterBuildResult:
    loaded = LoadedMasterConfig(
        template=template,
        master_path=template.config_path,
        master_doc={},
    )
    buckets: MutableMapping[Tuple[str, str, str], CandidateAccumulator] = OrderedDict()
    early_exits: MutableMapping[Tuple[str, str, str], EarlyExitAccumulator] = (
        OrderedDict()
    )
    metric_ids = []  # type: List[str]
    metric_styles = OrderedDict()  # type: MutableMapping[str, DynamicMetricStyle]

    for workbook_path in workbook_paths:
        ingest_node_workbook_dynamic(
            workbook_path=workbook_path,
            loaded=loaded,
            buckets=buckets,
            early_exits=early_exits,
            metric_ids=metric_ids,
            metric_styles=metric_styles,
        )

    default_null_rule = DEFAULT_NULL_RULE
    metric_headers = dynamic_metric_headers(
        metric_ids=metric_ids,
        metrics_by_id=template.metrics_by_id,
    )
    rows = []
    for bucket in buckets.values():
        row = OrderedDict()
        row["ID"] = resolve_early_exit_source_value(
            bucket=bucket,
            kind="roster_field",
            field_name="uid",
            default_null_rule=default_null_rule,
        )
        row["Last Name"] = resolve_early_exit_source_value(
            bucket=bucket,
            kind="roster_field",
            field_name="last",
            default_null_rule=default_null_rule,
        )
        row["First Name"] = resolve_early_exit_source_value(
            bucket=bucket,
            kind="roster_field",
            field_name="first",
            default_null_rule=default_null_rule,
        )
        row["Block"] = resolve_early_exit_source_value(
            bucket=bucket,
            kind="meta_field",
            field_name="block_number",
            default_null_rule=default_null_rule,
        )
        row["Fiscal Year"] = resolve_early_exit_source_value(
            bucket=bucket,
            kind="meta_field",
            field_name="fiscal_year",
            default_null_rule=default_null_rule,
        )
        row["Source"] = resolve_early_exit_source_value(
            bucket=bucket,
            kind="ingestion_metadata",
            field_name="source_workbook",
            default_null_rule=default_null_rule,
        )
        for metric_id, header in zip(metric_ids, metric_headers):
            row[header] = resolve_dynamic_metric_value(
                values=bucket.metric_values.get(metric_id, []),
                metric=template.metrics_by_id[metric_id],
            )
        rows.append(row)

    early_exit_rows = build_early_exit_rows(
        early_exits=early_exits,
        default_null_rule=default_null_rule,
    )
    return DynamicMasterBuildResult(
        rows=rows,
        early_exit_rows=early_exit_rows,
        metric_ids=metric_ids,
        metric_styles=metric_styles,
    )


def loaded_request_workbooks(loaded: LoadedMasterConfig) -> Sequence[Path]:
    workbook_paths = loaded.master_doc.get("_request_workbook_paths")
    if not isinstance(workbook_paths, list):
        raise ValueError("Internal master generation state is missing workbook paths")
    return [Path(value) for value in workbook_paths]


def attach_request_workbooks(
    loaded: LoadedMasterConfig, workbook_paths: Sequence[Path]
) -> LoadedMasterConfig:
    master_doc = dict(loaded.master_doc)
    master_doc["_request_workbook_paths"] = [str(path) for path in workbook_paths]
    return LoadedMasterConfig(
        template=loaded.template,
        master_path=loaded.master_path,
        master_doc=master_doc,
    )


def build_master_workbook(
    *,
    loaded: LoadedMasterConfig,
    rows: Sequence[OrderedDict[str, Any]],
    early_exit_rows: Sequence[OrderedDict[str, Any]] = (),
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER"

    headers = [
        column.get("header", column.get("column_id"))
        for column in loaded.master_doc["columns"]
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header))

    if early_exit_rows:
        early_exit_ws = wb.create_sheet("EARLY_EXITS")
        early_exit_headers = list(early_exit_rows[0].keys())
        for col_idx, header in enumerate(early_exit_headers, start=1):
            cell = early_exit_ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        for row_idx, row in enumerate(early_exit_rows, start=2):
            for col_idx, header in enumerate(early_exit_headers, start=1):
                early_exit_ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=row.get(header),
                )

    return wb


def build_dynamic_master_workbook(
    *,
    template: LoadedTemplateConfig,
    result: DynamicMasterBuildResult,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "MASTER"

    fixed_headers = [
        "ID",
        "Last Name",
        "First Name",
        "Block",
        "Fiscal Year",
        "Source",
    ]
    fixed_machine_headers = [
        "uid",
        "last",
        "first",
        "block_number",
        "fiscal_year",
        "source_workbook",
    ]
    metric_headers = dynamic_metric_headers(
        metric_ids=result.metric_ids,
        metrics_by_id=template.metrics_by_id,
    )
    headers = fixed_headers + metric_headers
    machine_headers = fixed_machine_headers + list(result.metric_ids)
    metric_start_col = len(fixed_headers) + 1
    data_start_row = 3
    data_end_row = max(data_start_row, len(result.rows) + data_start_row - 1)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        if col_idx >= metric_start_col:
            metric_id = result.metric_ids[col_idx - metric_start_col]
            style = result.metric_styles.get(metric_id)
            cell.fill = copy(style.header_fill) if style else HEADER_FILL
        else:
            cell.fill = HEADER_FILL

    for col_idx, header in enumerate(machine_headers, start=1):
        ws.cell(row=2, column=col_idx, value=header).font = HEADER_FONT
    ws.row_dimensions[2].hidden = True
    ws.freeze_panes = "G3"

    for row_idx, row in enumerate(result.rows, start=data_start_row):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header))
            if col_idx >= metric_start_col:
                metric_id = result.metric_ids[col_idx - metric_start_col]
                style = result.metric_styles.get(metric_id)
                if style and style.body_fill:
                    apply_table_cell_style(cell, copy(style.body_fill))

    for offset, metric_id in enumerate(result.metric_ids):
        col_idx = metric_start_col + offset
        col_letter = get_column_letter(col_idx)
        metric = template.metrics_by_id[metric_id]
        if metric.get("type") == "timed":
            for row_idx in range(data_start_row, data_end_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0"
        elif is_numeric_metric(metric):
            for row_idx in range(data_start_row, data_end_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0.00"
        add_min_pass_highlight_rule(
            ws=ws,
            metric=metric,
            first_cell="{0}{1}".format(col_letter, data_start_row),
            cell_range="{0}{1}:{0}{2}".format(
                col_letter,
                data_start_row,
                data_end_row,
            ),
        )

    if result.early_exit_rows:
        early_exit_ws = wb.create_sheet("EARLY_EXITS")
        early_exit_headers = list(result.early_exit_rows[0].keys())
        for col_idx, header in enumerate(early_exit_headers, start=1):
            cell = early_exit_ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        for row_idx, row in enumerate(result.early_exit_rows, start=2):
            for col_idx, header in enumerate(early_exit_headers, start=1):
                early_exit_ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=row.get(header),
                )

    return wb


def generate_master_workbook(request: MasterGenerationRequest) -> Path:
    if request.dynamic:
        template = load_generation_inputs(
            TemplateGenerationRequest(
                config_path=request.config_path,
                entry_rows=1,
            )
        )
        output_path = (
            request.output_path.resolve()
            if request.output_path
            else default_dynamic_output_path(template)
        )
        batch = resolve_master_workbook_batch(request, output_path=output_path)
        result = build_dynamic_master_result(
            template=template,
            workbook_paths=batch.workbook_paths,
        )
        workbook = build_dynamic_master_workbook(
            template=template,
            result=result,
        )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        move_processed_workbooks(batch)
        return output_path

    base_loaded = load_master_generation_inputs(request)
    output_path = (
        request.output_path.resolve()
        if request.output_path
        else default_output_path(base_loaded)
    )
    batch = resolve_master_workbook_batch(request, output_path=output_path)
    loaded = attach_request_workbooks(base_loaded, workbook_paths=batch.workbook_paths)
    result = build_master_result(loaded)
    workbook = build_master_workbook(
        loaded=loaded,
        rows=result.rows,
        early_exit_rows=result.early_exit_rows,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    move_processed_workbooks(batch)
    return output_path


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    output_path = generate_master_workbook(request_from_namespace(args))
    print("Master workbook generated: {0}".format(output_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
